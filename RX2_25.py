import tkinter as tk
from tkinter import ttk,messagebox
import os
import sys

import time
from datetime import date
from datetime import datetime,timedelta

from tkinter import *
import tkinter.font as font
import base64
os.environ['CUDA_VISIBLE_DEVICES'] = '0'
import pyautogui as pg
import numpy as np
import cv2
from PIL import ImageTk, Image
from openpyxl import Workbook
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from easygui import *
import easygui
import win32api,win32con
import codecs
import collections
import math
import dicom as dcm
import sqlite3 as sq
import glob
import shutil

import pywinauto
from pywinauto import application

import json




def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.getcwd()
    return os.path.join(base_path, relative_path)

####################################################################################################################################################################################################################################################################################################################################################################################################################
####################################################################################################################################################################################################################################################################################################################################################################################################################
class load_path:
    def load_element(self, element):
        self.jsn_list="projet.json"
        try:
          openfile = open(self.jsn_list, 'r')
          jsn_conten = json.load(openfile)
        except:
            outfile = open(self.jsn_list, "w")
            if self.jsn_list=="projet.json":
                J_dict={"projet":"Rgz-2",
                        "Nuance":"x70  MPSL 2",
                        "Diameter":"1016 mm",
                        "Epaisseur":"12,70 mm",
##                        "ip_address":"192.168.2.4",
####                    "values":[-60,-30,-20,-15,-10,0,10,15,20,30,60,80],
##                        "values_wait":25,
##                       
##                        
##                        "defaut_AREP":["AA","BA","BU","DL","F","SCVE"],
##                        "defaut_AMEULER":["AA","BA","BU","DL","F","SCVE"],
##                        "DEFAULT_ACH":['AA/D','BA/D','BU/D','DL/D','CF/D','MM/D','AA/F','BA/F','BU/F','DL/F','CF/F','MM/F'],
##                        "defaut_OK":["AA","BA","BU","DL","F","SCVE"],
##                        
##                        
##                        "Classes":["AA_BA","DL","BU","AN,MMC","F","WELD_IMPERFECTION"],
                        "LISTE_DES_POST":[5,13,13,21,21,5,4],
                        
                        "LISTE_operateurs":['BOUZID YASSINE','BOUZID YASSINE','BOUZID YASSINE'],
                        "LISTE_Projets":['Rgz-2','CEEG KD/AL'],
                        
                        "PATH_PROJET": r"C:\Users\111\Desktop\1750",
##                        "PATH_CLIENT": r"C:\Users\111\Desktop\client",
##                        "PATH_SCOPIE": r"//YASSINE-PC/shared folder/SCOPY",
##                        "path_fab"   : r"C:\Users\YASSINE\Desktop\SCRIPTS\fabrication",
##                        "PATH_prog"  : r"C:\Program Files\YXLON\Y.Image 3500\IMAGE_3500.exe",
##                        "top": 97,
##                        "left": 491,
##                        "width": 1000,
##                        "height": 1000,        
##                        #"bits":8,
                        "integration_time":9
                        }   
            json_object = json.dumps(J_dict, indent=2)
            outfile.write(json_object)
            jsn_conten = J_dict
        return jsn_conten[element]
    
P = load_path() 

##def load_list(list_of_defcts):
##    try:
##        text_default=open(list_of_defcts,'r')
##        content=(text_default.read()).split("\n")
##        text_default.close()
##        print("content=",content)
##    except:
##        text_default=open(list_of_defcts,'w')
##        if list_of_defcts=="defaut_AREP.txt":
##            text_default.write("AA\nBA\nBU\nDL\nF\nSCVE\n")
##        if list_of_defcts=="defaut_AMEULER.txt":
##            text_default.write("AA\nBA\nBU\nDL\nF\nSCVE\n")
##        if list_of_defcts=="defaut_ACHUTE.txt":
##            text_default.write("AA\nBA\nBU\nDL\nF\nSCVE\n")
##        if list_of_defcts=="defaut_OK.txt":
##            text_default.write("AA\nBA\nBU\nDL\nF\nSCVE\n")
##            
##        if list_of_defcts=="PATH_CLIENT.txt":
##            text_default.write(r"C:\Users\YXLON.YXLON-PC\Desktop\GZ-2 CLIENT SONATRACH")
##            
##        if list_of_defcts=="LISTE_operateurs.txt":
##            text_default.write('BOUZID YASSINE\nBOUZID YASSINE\nBOUZID YASSINE')
##            
##        if list_of_defcts=="LISTE_Projets.txt":
##            text_default.write('CEEG KD/AL\nR-GZ2')
##            
##            
##        if list_of_defcts=="PATH_PROJET.txt":
##            text_default.write(r"Z:\GZ-2")
##            
##        if list_of_defcts=="PATH_app.txt":
##            text_default.write(r"C:\Program Files\YXLON\Y.Image 3500\IMAGE_3500.exe")
##            
##        if list_of_defcts=="LISTE_DES_POST.txt":
##            text_default.write('05\n13\n13\n21\n21\n05\n04')
##            
##        if list_of_defcts=="PATH_CN_DB.txt":
##            text_default.write(r"C:\Users\111\Desktop\REPORT APPS\CN\CN_APP_16\RAPPORT_CN.db")
##
##                     
##        text_default.close()
##        text_default=open(list_of_defcts,'r')
##        content=(text_default.read()).split("\n")
##        text_default.close()
##    return content

def covert_to_excel(listing):
    listing =  (str(listing)).replace('[','')
    listing =  listing.replace(']','')
    listing =  listing.replace("'","")
    listing =  listing.replace("\n","")
    return listing
def covert_PATH(listing):
    listing =  (str(listing)).replace('/','_')
    listing =  listing.replace("\\", '_')
    listing =  listing.replace("<","")
    listing =  listing.replace(">","")                           
    listing =  listing.replace("?","")
    listing =  listing.replace(":","")
    listing =  listing.replace("*","")
    listing =  listing.replace('"',"")
    listing =  listing.replace("|","")
    return listing


## DECLARATION OF DATABASE COLUMN NAMES:

#NAME_DATABASE=  covert_to_excel(load_list("PATH_CN_DB.txt"))
NAME_DATABASE= P.load_element('projet')#PROJECT#"RAPPORT_RX2"#codecs.decode(f"{NAME_DATABASE}",'unicode_escape')
NAME_TABLE = "RAPPORT_RX2"
NAME_REPORT= "REPORT"
col1="TUBE"
col2="N°INTEGRATION"
col3="DESIGNATIONS"
col4="ZIP_IMAGES"

####################################################################################################################################################################################################################################################################################################################################################################################################################
####################################################################################################################################################################################################################################################################################################################################################################################################################

i=int()
i=0
j=int()
j=1
k=int()
k=1
j_k = int()
j_k = 1
max_line = 21
result_list=[]
iqi_not_checked = True


timing = datetime.now()
if 21<=timing.hour<24 :
    time_value=datetime.now()+timedelta(1)
else:
    time_value=datetime.now()


time_value=time_value.strftime("%d-%m-%y")
print(time_value)

output = ""
defect_name1,defect_number1,defect_letter1,defect_FAR1,SOUNDAGE1= "","","","",""

####################################################################################################################################################################################################################################################################################################################################################################################################################
########################################################################---------------------PYDICOME READER-------------------------#################################################################################################################################################################################################################################################################################################################


# TO DO: FINISH STORING IN DATABSE


def creating_the_zip_file():
    dir_name="zip folder"
    output_filename="Images"
    if not os.path.exists(dir_name):
            os.makedirs(dir_name) 
    list_of_filmes= glob.glob(newpath+'/*.dcm')
    for i, each_film in enumerate(list_of_filmes):
        img = dcm.read_file(each_film)            #np.median(img.pixel_array)+50-img.pixel_array)
        cv2.imwrite(f"{dir_name}\{getname}_{i}.jpg", np.median(img.pixel_array)+50-img.pixel_array)
        #shutil.copy(each_film, dir_name)
    shutil.make_archive(output_filename, 'zip', dir_name)
    
    for each_jpg in glob.glob(dir_name+'/*.jpg'):
        os.remove(each_jpg)

def convertToBinaryData(filename):
    # Convert digital data to binary format
    with open(filename, 'rb') as file:
        blobData = file.read()
    return blobData

def create_atable(NAME_DATABASE,NAME_TABLE,NAME_REPORT,col1,col2,col3,col4):
    cnn =sq.connect(f'{NAME_DATABASE}.db')
    c = cnn.cursor()
    c.execute(f"CREATE TABLE IF NOT EXISTS {NAME_TABLE} ({NAME_REPORT} text, {col1} text, {col2}  integer, {col3}  text, {col4} blob)")
    cnn.commit()
    cnn.close()
    print("done!")

def insert_one_record(NAME_DATABASE,NAME_TABLE,one_record):
    cnn =sq.connect(f'{NAME_DATABASE}.db')
    c = cnn.cursor()
    c.execute(f"INSERT INTO {NAME_TABLE} VALUES(?,?,?,?,?)",one_record)
    cnn.commit()
    cnn.close()
    print("done!")
    
def select_by_pipe_name(NAME_DATABASE,NAME_TABLE, PIPENAME):# done
    
    cnn =sq.connect(f'{NAME_DATABASE}.db')
    c = cnn.cursor()
    c.execute(f"SELECT * FROM {NAME_TABLE} WHERE TUBE = '{PIPENAME}'")

    rows = c.fetchall()

    for row in rows:
        print(row[:3])
    return rows




####################################################################################################################################################################################################################################################################################################################################################################################################################
#######################
         
def save_file():
    global getname
    pg.moveTo(100,600)
    pg.click()
    pg.keyDown('ctrl')
    pg.press('s')# SAVE
    pg.keyUp('ctrl')
    #time.sleep(int(delay_save.get()))
    time.sleep(float(l_variable.get()))
    
    try:
            #the_x,the_y= pg.locateCenterOnScreen('name.png', grayscale=True,confidence = confidenceE.get())
            x, y = pg.locateCenterOnScreen('name.jpg', grayscale=True,confidence =.9)
            pg.moveTo(x +200, y)
            pg.click()
            #print(x +200, y)
    except Exception as e:
        print("the exception is", e)
        pg.moveTo(330,855)
        pg.click()
    file1 = "{}\{}".format(newpath,getname+"-{}".format(output))
    
    caps_status = win32api.GetKeyState(win32con.VK_CAPITAL)
    if caps_status==0:
        #print('CapsLock is off')
        file = file1[2:]
        pg.press('capslock')
        pg.press('capslock')
        pg.press('capslock')
        pg.write(str(file1[:2]))# SAVE
        pg.press('capslock')
        pg.write(file)
        #print("SAVED file======== ", file)
        pg.press('enter')
        pg.press('capslock')
    else:
        #print('CapsLock is on')
        pg.press('capslock')
        pg.press('capslock')
        file = file1[2:]
        pg.write(str(file1[:2]))# SAVE
        pg.press('capslock')
        pg.write(file)
        #print("SAVED file======== ", file)
        pg.press('enter')
        pg.press('capslock')
    
def tube_finished(event):
    global i,j,k,result_list,ws,operators_names1,getname,j_k,NAME_DATABASE,NAME_TABLE,NAME_REPORT,col1,col2,col3,col4,pipe_name
   # path = covert_to_excel(load_list("PATH_PROJET.txt"))
   # path = codecs.decode(f"{path}",'unicode_escape')
   #j_project =  load_json("projet.json")["Nuance"]+ "\n"+load_json("projet.json")["Diameter"]+"\n"+load_json("projet.json")["Epaisseur"]
    if not os.path.isfile(r"{}\RAPPORT RX2 N°{}_{}_{}.xlsx".format(path,k,POST1,time_value)):
        wb = load_workbook('RX2_template.xlsx')
        wb.save(r"{}\RAPPORT RX2 N°{}_{}_{}.xlsx".format(path,k,POST1,time_value))
    if i ==0:
        print("at leatst one integration")
        devlabel.config(text="MINIMUM UNE INTERATION",bg =color)
        return
    j_project =  P.load_element("Nuance")+ "\n"+P.load_element("Diameter")+"\n"+P.load_element("Epaisseur")
    
    wb = load_workbook(r"{}\RAPPORT RX2 N°{}_{}_{}.xlsx".format(path,k,POST1,time_value))    
    ws = wb.active
    ws['G1'] = f"Page: {k}"
    ws['A5'] = f"Projet: {PROJECT}"
    ws['F4'] =  j_project
    
    ws['G1'].font = Font(size=18)
    #ws['D2'] = f'Rapport de contrôle\n RX2 NUMERIQUE N° "{k}"'
    # set the "Equipe" and the "Post" from the form
    #EQUIPE1 POST1 operators2_names1 operators_names1
    ws['A5'] =f'Equipe:   "{EQUIPE1}".\nPost:       "{POST1}".'
    # set the name of the operators
    
    
    ws['A3'] = f'DATE: {time.strftime("%d-%m-%y")}'
    ws['A3'].font = Font(size=18)
    ws['A{}'.format(j+10)] = j
    ws['B{}'.format(j+10)] = getname
    ws['C{}'.format(j+10)] = i
    
    if pipe_name != "IQI":
        db_result_list=[]
        list_of_filmes= os.listdir(newpath)
        for i, each_film in enumerate(list_of_filmes):
            if each_film.endswith(".dcm"):
                print(each_film[6:-8])
                db_result_list.append(each_film[6:-8])
    else:
        db_result_list=[f"F-{time_value}-{EQUIPE1}-{POST1}".replace(" ", "")]

        
    xl_results0 =  (str(result_list)).replace('[','')
    xl_results1 =  xl_results0.replace(']','')
    xl_results2 =  xl_results1.replace("'","")
    # specifiying the width of the celule
    x = int(len(xl_results2)/49)
    #print("xl_results2=",xl_results2)
    #print("x=",x)
    if x==0:
        ws.row_dimensions[j+10].height = 20
        ws['D{}'.format(j+10)] =xl_results2
    else:
        ws.row_dimensions[j+10].height = (x+1)*17
        j_k+=x
        ws['D{}'.format(j+10)] =xl_results2
    
    ws['A32'] =f'Nom et Prénom :\n{operators_names1}\n Visa:'
    ws['E32'] =f'                                                     Nom et Prénom :\n                                                     {operators2_names1}\n                                                     Visa:'

    #CREATE A TABLE FOR THE REPORT
    #TABLE COLUMNS
    
##    col5="ACTION"
##    col6="ZIP_IMAGES"
##    
    create_atable(NAME_DATABASE,NAME_TABLE,NAME_REPORT,col1,col2,col3,col4)
    print("table created!")
    

    #INSERT A RECORD INTO THE TABLE%%%%%%%%%%%%%%%%%%%% DATABASE %%%%%%%%%%%%
    creating_the_zip_file()
    
    # RECORD DECLARATIONS
    NAME_REPORT_CURRENT="{}_RAPPORT_RX2 N°{}_{}_{}_AND_{}_{}".format(PROJECT,k,POST1,operators_names1,operators2_names1,time_value)
    REC1=PIPE_NAME.get().upper()
    REC2= i
    REC3= covert_to_excel(db_result_list)
    REC4= convertToBinaryData("Images.zip")
    
    one_record=(NAME_REPORT_CURRENT,REC1,REC2,REC3,REC4)
    
    print("pipe name is =================================================",str(PIPE_NAME.get()).upper())
    if str(PIPE_NAME.get()).upper()=="IQI":
        pass
    else:
        insert_one_record(NAME_DATABASE,NAME_TABLE,one_record)
    
        print("record inserted!")
    
                                             
    wb.save(r"{}\RAPPORT RX2 N°{}_{}_{}.xlsx".format(path,k,POST1,time_value))
    i=0
    result_list=[]
    j+=1
    DEFECT_NUMB_lab.config(text=i)
    finish_tube.config(text="Tube N°{}".format(j),bg ="RoyalBlue1")
    
    if j>= max_line:
        win32api.ShellExecute(
        0,
        "print",
        r"{}\RAPPORT RX2 N°{}_{}_{}.xlsx".format(path,k,POST1,time_value),
        None,
        ".",
        0
        )
        #os.startfile(r"{}\POST REPORT N°{} at {}.xlsx".format(path,k,time_value), 'print')
        report_closed.config(text="P{}-PRINTED".format(k),bg = "green2")
        if not os.path.exists(r"{}\RAPPORTS\RX2".format(path)):
            os.makedirs(r"{}\RAPPORTS\RX2".format(path)) 
        shutil.move(r"{}\RAPPORT RX2 N°{}_{}_{}.xlsx".format(path,k,POST1,time_value),r"{}\RAPPORTS\RX2".format(path))
        finish_tube.config(text="INSÈRE",bg =btncolor)
        k+=1
        j=1
        j_k=1
        result_list=[]
        
    #rest all widget
    PIPE_NAME.config(state='normal')
    PIPE_NAME.delete(0,"end")
    defect_name.set("")
    defect_number.set("")
    defect_letter.set("")
    defect_FAR.set("") 
    print("!!! tube finished !!!")
    PIPE_NAME.focus_set()
    
    
    
def report_closed_func(event):
    global k,i,j
    
   # path = covert_to_excel(load_list("PATH_PROJET.txt"))
   # path = codecs.decode(f"{path}",'unicode_escape')
    
    confirmation = messagebox.askquestion("CONFIRMATION CLOTURAGE!","VOULEZ VRAIMENT CLOTURE LE RAPPORT?")
    if j ==1:
        print("at leatst one TUBE")
        devlabel.config(text="RAPPORT VIDE !!",bg ="yellow")
        
    if not os.path.isfile(r"{}\RAPPORT RX2 N°{}_{}_{}.xlsx".format(path,k,POST1,time_value)):
        j_project =  P.load_element("Nuance")+ "\n"+P.load_element("Diameter")+"\n"+P.load_element("Epaisseur")
        wb = load_workbook('RX2_template.xlsx')
        ws = wb.active
        ws['G1'] = f"Page: {k}"
        ws['A5'] = f"Projet: {PROJECT}"
        ws['G1'].font = Font(size=18)
        ws['A5'] =f'Equipe:   "{EQUIPE1}".\nPost:       "{POST1}".'
        ws['A3'] = f'DATE: {time.strftime("%d-%m-%y")}'
        ws['A3'].font = Font(size=18)
        ws['A3'] = f'DATE: {time.strftime("%d-%m-%y")}'
        ws['F4'] =  j_project
        ws['A3'].font = Font(size=18)
        ws['A32'] =f'Nom et Prénom :\n{operators_names1}\n Visa:'
        ws['E32'] =f'                                                     Nom et Prénom :\n                                                     {operators2_names1}\n                                                     Visa:'                            

        wb.save(r"{}\RAPPORT RX2 N°{}_{}_{}.xlsx".format(path,k,POST1,time_value))
        
        
    if confirmation=="yes":   

        # PRINT THE REPPORT:
        win32api.ShellExecute(
        0,
        "print",
        r"{}\RAPPORT RX2 N°{}_{}_{}.xlsx".format(path,k, POST1,time_value),
        None,
        ".",
        0
        )
        #rest all widget
        #os.startfile(r"{}\POST REPORT N°{} at {}.xlsx".format(path,k,time_value), 'print')
        PIPE_NAME.delete(0,"end")
        defect_name.set("")
        defect_number.set("")
        defect_letter.set("")
        defect_FAR.set("")
        print("!!!!!!!!rapport closed!!!!!!!!")
        devlabel.config(text= "Rapport closed!",fg="orange",bg="yellow")
        
        report_closed.config(text="P{}-PRINTED".format(k),bg = "green2")
        if not os.path.exists(r"{}\RAPPORTS\RX2".format(path)):
            os.makedirs(r"{}\RAPPORTS\RX2".format(path)) 
        shutil.move(r"{}\RAPPORT RX2 N°{}_{}_{}.xlsx".format(path,k,POST1,time_value),r"{}\RAPPORTS\RX2".format(path))
        finish_tube.config(text="INSÈRE",bg =btncolor)
        k+=1
        i=0
        j=1
        j_k=1
        PIPE_NAME.focus_set()


def open_and_create_folder(event):
    
    global i,newpath,path,wb,result_list,output,defect_name1,defect_number1,defect_letter1,defect_FAR1,getname
    pipe_name=str(PIPE_NAME.get()).upper()
    
   # path = covert_to_excel(load_list("PATH_PROJET.txt"))
   # path = codecs.decode(f"{path}",'unicode_escape')

    if PIPE_NAME.get()=="" or (pipe_name[0] not in ['A','B','C','D','E','I']):
        print("File up pipe name, path entries and defect name ")
        devlabel.config(text= "File up pipe name,\n path entries and defect name!!",fg="red",bg="yellow")
        return
    
    if len(pipe_name)==2:
        pipe_name=str(PIPE_NAME.get()).upper()
        print("length====",len(pipe_name),pipe_name[1])
        #A0001
        getname = (pipe_name[0]+"0"+"0"+"0"+pipe_name[1]).upper()
        name  = str(getname[0])
        newpath = "{}\{}\{}".format(path,name,getname)
        print("getname",getname)
        
    elif len(PIPE_NAME.get())==3:
        pipe_name=str(PIPE_NAME.get()).upper()
        #A0012
        if pipe_name == "IQI":
            getname = pipe_name
            newpath = "{}\{}".format(path,getname)
            
        else:
            getname = (pipe_name[0]+"0"+"0"+pipe_name[1]+pipe_name[2]).upper()
            name  = str(getname[0])
            newpath = "{}\{}\{}".format(path,name,getname)
            
    elif len(PIPE_NAME.get())==4:
        pipe_name=str(PIPE_NAME.get()).upper()
        #A0123
        getname = (pipe_name[0]+"0"+pipe_name[1]+pipe_name[2]+pipe_name[3]).upper()
        name  = str(getname[0])
        newpath = "{}\{}\{}".format(path,name,getname)
        
    elif len(PIPE_NAME.get())==5:
        pipe_name=str(PIPE_NAME.get()).upper()
        #A1234
        devlabel.config(text= "DEVELOPED BY BOUZID YASSINE \n CND-INSPECTOR RT-II 2020",fg="black",bg=color)
        getname = str(PIPE_NAME.get()).upper()
        name  = str(getname[0])
        newpath = "{}\{}\{}".format(path,name,getname)
        print("getname= ",getname)
        
    elif ((PIPE_NAME.get()).upper().find('-BIS') != -1):
        pipe_name=str(PIPE_NAME.get()).upper()
        devlabel.config(text= "BIS TUBE!!!",fg="black",bg="yellow")
        getname = str(PIPE_NAME.get()).upper()
        name  = str(getname[0])
        newpath = "{}\{}\{}".format(path,name,getname)
        
    else:
        devlabel.config(text= "UNCORRECT TUBE NUMBER!",fg="black",bg=color)
        return



        
   
    if i==0 and os.path.exists(newpath):
        print("file exists!!!XXXXXXXXXXXXX!!")
        devlabel.config(text= "Folder exists!!!",fg="orange",bg="yellow")
        
    
    if not os.path.exists(newpath):
        os.makedirs(newpath)
        
    os.startfile(newpath)
    PIPE_NAME.delete(0,"end")
    PIPE_NAME.insert(0,getname)
    ROW = select_by_pipe_name(NAME_DATABASE,NAME_TABLE,PIPE_NAME.get().upper())
    if ROW:
        print("messagebox this file is existed")
        messagebox.showinfo("FOLDER EXISTES!",f"TUBE UPDATE N°{len(ROW)}: \n{[R[1:-1] for R in ROW]}")

def manque_iqi(test_check):
    global defect_number1
    messagebox.showinfo("INDICATEUR DE QUALITE D'IMAGE!",f"MANQUE IQI: {test_check}")
    PIPE_NAME.delete(0,"end")
    PIPE_NAME.insert(0,'IQI')
    defect_name1,defect_number1,SOUNDAGE1,defect_letter1,defect_FAR1= "","","","",""
    defect_name.set("")
    SOUNDAGE.set("")
    defect_number.set("")
    defect_letter.set("")
    defect_FAR.set("")
    # deselect the checkbox rm1
    RM1.deselect()
    RM1.config(bg =color)
    PIPE_NAME.config(state='disabled')  
    
test_check = ""
      
def integration(event):
    global i,newpath,pipe_name,path,wb,result_list,output,defect_name1,defect_number1,defect_letter1,defect_FAR1,SOUNDAGE1,getname,iqi_not_checked,test_check
    pipe_name=str(PIPE_NAME.get()).upper()
   # path = covert_to_excel(load_list("PATH_PROJET.txt"))
   # path = codecs.decode(f"{path}",'unicode_escape')

    timing = datetime.now()
    
    

    if PIPE_NAME.get()=="" or (defect_name1=="" and PIPE_NAME.get().upper()!="IQI") or (pipe_name[0] not in ['A','B','C','D','E','I']):
        print("File up pipe name, path entries and defect name ")
        devlabel.config(text= "File up pipe name,\n path entries and defect name!!",fg="red",bg="yellow")
        PIPE_NAME.config(state='normal')
        return

    if len(pipe_name)==2:
        pipe_name=str(PIPE_NAME.get()).upper()
        print("length====",len(pipe_name),pipe_name[1])
        #A0001
        getname = (pipe_name[0]+"0"+"0"+"0"+pipe_name[1]).upper()
        name  = str(getname[0])
        newpath = "{}\{}\{}".format(path,name,getname)
        print("getname",getname)
        
    elif len(PIPE_NAME.get())==3:
        pipe_name=str(PIPE_NAME.get()).upper()
        #A0012
        if pipe_name == "IQI":
            getname = pipe_name
            defect_name.set("")
            defect_number.set("")
            defect_letter.set("")
            defect_FAR.set("")
            newpath = "{}\{}".format(path,getname)
            
        else:
            getname = (pipe_name[0]+"0"+"0"+pipe_name[1]+pipe_name[2]).upper()
            name  = str(getname[0])
            newpath = "{}\{}\{}".format(path,name,getname)
            
    elif len(PIPE_NAME.get())==4:
        pipe_name=str(PIPE_NAME.get()).upper()
        #A0123
        getname = (pipe_name[0]+"0"+pipe_name[1]+pipe_name[2]+pipe_name[3]).upper()
        name  = str(getname[0])
        newpath = "{}\{}\{}".format(path,name,getname)
        
    elif len(PIPE_NAME.get())==5:
        pipe_name=str(PIPE_NAME.get()).upper()
        #A1234
        devlabel.config(text= "DEVELOPED BY BOUZID YASSINE \n CND-INSPECTOR RT-II 2020",fg="black",bg=color)
        getname = str(PIPE_NAME.get()).upper()
        name  = str(getname[0])
        newpath = "{}\{}\{}".format(path,name,getname)
        print("getname= ",getname)
        
    elif ((PIPE_NAME.get()).upper().find('-BIS') != -1):
        pipe_name=str(PIPE_NAME.get()).upper()
        devlabel.config(text= "BIS TUBE!!!",fg="black",bg="yellow")
        getname = str(PIPE_NAME.get()).upper()
        name  = str(getname[0])
        newpath = "{}\{}\{}".format(path,name,getname)
        
    else:
        devlabel.config(text= "UNCORRECT TUBE NUMBER!",fg="black",bg=color)
        PIPE_NAME.config(state='normal')
        return
    
    if iqi_check <= timing.hour and iqi_not_checked and pipe_name != "IQI":
        test_check ="test-2"
        manque_iqi(test_check)
        iqi_not_checked = False
        return "break"
    
    if i==0 and j==1 and iqi_check > timing.hour and k==1 and pipe_name != "IQI":
        test_check ="test-1"
        manque_iqi(test_check)
        return "break"
    
    print("iqi_check === timing.hour iqi_not_checked",iqi_check,timing.hour,iqi_not_checked)    
        

        
    #name  = str(getname[0])
    
    #newpath = "{}\{}\{}".format(path,name,getname)
    PIPE_NAME.delete(0,"end")
    PIPE_NAME.insert(0,getname)
    PIPE_NAME.config(state='disabled')        
    if i==0 and os.path.exists(newpath):
        print("file exists!!!XXXXXXXXXXXXX!!")
        devlabel.config(text= "Folder exists!!!",fg="orange",bg="yellow")
    if not os.path.exists(newpath):
        os.makedirs(newpath)
    i+=1
    
##    time_integration = int("9")
##    print("Integration started!")
##    print("i=",i)

    #time.sleep(time_integration)
    pg.moveTo(120,120)
    pg.click()
    pg.press('f4')# INTEGRATE
    pg.moveTo(167,55)
    pg.click()# CONFIRM INTEGRATION
    #print("defect number=",defect_name1)
    
    #time.sleep(9)
    time.sleep((P.load_element("integration_time")))
    if var2.get() == 1:
        
            
        output = "{}{}{}{}RM1{}".format(defect_name1,defect_number1,defect_letter1,SOUNDAGE1,defect_FAR1)#f'{defect_name.get()}{defect_number.get()}{defect_letter.get()}RM1{defect_FAR.get()}'
        print("output=",output)
    else:
        if pipe_name == "IQI":
            output =f"F-{time_value}-{EQUIPE1}-{POST1}-{test_check}".replace(" ", "")
        else:
            output = "{}{}{}{}{}".format(defect_name1,defect_number1,defect_letter1,SOUNDAGE1,defect_FAR1)
    if output:
        print("output=",output)
        result_list.append(output)
        
        defect_name1,defect_number1,SOUNDAGE1,defect_letter1,defect_FAR1= "","","","",""
        defect_name.set("")
        SOUNDAGE.set("")
        defect_number.set("")
        defect_letter.set("")
        defect_FAR.set("")
        # deselect the checkbox rm1
        RM1.deselect()
        RM1.config(bg =color)

    save_file()
    """
    pg.moveTo(120,55)# PRESS LIVE
    pg.click()
    pg.moveTo(120,55)# PRESS LIVE
    pg.click()
    """
    #time.sleep(int(DELAYWORKER.get()))

    DEFECT_NUMB_lab.config(text = " {}".format(i), bg ="yellow")
##        time_delay = int("0")
##        print("time delay = ",time_delay)
##        time.sleep(time_delay)
    print('finished')
    defect_name.focus_set()
    devlabel.config(text= "DEVELOPED BY BOUZID YASSINE \n CND-INSPECTOR RT-II 2020",fg="black",bg=color)
    


def check5(event):
    global j
    if var7.get()!=1:
        j_variable_lab.grid(row = 7 , column= 0, padx = 10,sticky="W")
        j_variable.grid(row = 7 , column= 1, padx = 40,pady = 5,sticky="W")
        j_variable_btn.grid(row =7,column=2, padx = 15,pady = 5,sticky="E")
        k_variable_lab.grid(row = 8 , column= 0, padx = 10,pady = 5,sticky="W")
        k_variable.grid(row = 8 , column= 1, padx = 40,pady = 5,sticky="W")
        #j = int(j_variable.get())
        #print("var7=", var7.get(),"j=",j)
        root.geometry("410x470")
                
    else:
        j_variable.grid_forget()
        j_variable_lab.grid_forget()
        j_variable_btn.grid_forget()
        k_variable_lab.grid_forget()
        k_variable.grid_forget()
        
        #print("var7=", var7.get(),"j=",j)
        root.geometry("410x400")


def j_variablefunc():
    global j,k
    if var7.get()==1:
        j = int(j_variable.get())
        k = int(k_variable.get())
        j_variable_btn.config(text=f"R N°{k}/L N°{j}",bg="green2")
        print("var7=", var7.get(),"j,k=",j,k)

    
def quitt():
    root.quit()
    sys.exit()
    #top = Toplevel()

def show_frame(frame):
    global path
    if str(operators_names.get())=="" or  str(operators2_names.get())=="" or str(POST.get())=="" or str(Projet.get())=="":
        operator1_names_lab.config(bg = "orange1")
        operator2_names_lab.config(bg = "orange1")
        POST_lab.config(bg = "orange1")
        PROJECT_lab.config(bg = "orange1")
        return
##    else:
##        path_app = covert_to_excel(load_list("PATH_app.txt"))
##        path_app = codecs.decode(f"{path_app}",'unicode_escape')
##        print("path_app=",path_app)
####        try:
##            print("this is the password you wrot =",f"{operators_names.get()} ",f"{PASSWORD_OP.get()}")
##            app = application.Application()
##            
##            app.start(path_app)#r"C:\Program Files\YXLON\Y.Image 3500\IMAGE_3500.exe")
##            time.sleep(.5)
##            app.IMAGE3500Etablissementdelaconnexion.ComboBox.Select(f"{operators_names.get()}")
##            app.IMAGE3500Etablissementdelaconnexion["enter votre mot de passe:Edit"].type_keys(f"{PASSWORD_OP.get()}")
##            #arret
##            #app.IMAGE3500Etablissementdelaconnexion.Button2.click()***
##            #ok
##            app.IMAGE3500Etablissementdelaconnexion.Ok.click()
##            #app.IMAGE3500DD.Ok.click()
##            PASSWORD_OP.delete(0,"end")
##        except Exception as e :print("no programe is found",e)
        
##        path = covert_to_excel(load_list("PATH_PROJET.txt"))
##        path = codecs.decode(f"{path}",'unicode_escape')
##        path = f"{path}\{covert_PATH(Projet.get())}"
       
        
            
    
    if frame == PIPE_FRAME_FRAME:
        frame.tkraise()
        frame.grid(row =0,column=0,sticky='nsew')
        root.geometry("410x400+1500+90")
        path = P.load_element("PATH_PROJET") +f'\{PROJECT}'
        try:
            if not os.path.exists(path):
                    os.makedirs(path) 
            print(path)
        except Exception as e:
            messagebox.showinfo("NETWORK PROBLEM!","CONNECTION PROBLEM VERIFY LAN CABLE\n Problème réseau, vérifier le cable réseau local ")
            print(e)
            return
    else:
        frame.tkraise()
        frame.grid(row =0,column=0,sticky='nsew')
        root.geometry("330x260+1500+90")




root = Tk()
root.config(bg ="white")
root.focus_force()
root.rowconfigure(0,weight= 1)
root.columnconfigure(0,weight=1)

color = "light sky blue"
color = "deep sky blue"
btncolor ="gold"



starting_FRAME = tk.Frame(root, width=100, height=100, background=color)
starting_FRAME.grid(row =0,column=0,sticky='nsew')

PIPE_FRAME_FRAME = tk.Frame(root, width=50, height=50, background=color )
#PIPE_FRAME_FRAME.pack(fill= "both", expand ="YES" , pady = 0, padx=10)
#PIPE_FRAME_FRAME.config(width=500, bg ="white")

starting = LabelFrame(starting_FRAME, text = "INFO",  width = 35,height =20,font =("Helvetica",10,"bold"),bg =color )
starting.grid(row =0,column=0,padx=2)

#PIPE_FRAME = LabelFrame(PIPE_FRAME_FRAME, text = "INFO",  width = 35,height =20,font =("Helvetica",10,"bold"),bg =color )
#PIPE_FRAME.grid(row =0,column=0,sticky='nsew')

PIPE_ET_DESIGNATION = LabelFrame(PIPE_FRAME_FRAME, text = "PIPE ET DESIGNATION:",  width = 35,height =20,font =("Helvetica",10,"bold"),bg =color,labelanchor = "n")
PIPE_ET_DESIGNATION.grid(row =0,column=0, columnspan=4,padx= 10,pady = 10)

#ACTIONS_FRAME = LabelFrame(PIPE_FRAME_FRAME, text = "ACTIONS:",  width = 35,height =20,font =("Helvetica",10,"bold"),bg =color )
#ACTIONS_FRAME.grid(row =2,column=0,sticky='nsew', columnspan=4)




    
################################################################################################ STARTING FRAME ##########################################################################################

#OPERATOR1 LABEL

operator1_names_lab = Label(starting, text= "OPERATEUR 01:",font =("Helvetica",10,"bold"), bg =color)
operator1_names_lab.grid(row = 0 , column= 0, padx = 10,sticky="W", pady=5)



# combobox5
def focus_func(event):
    global operators_names1
    if str(operators_names.get())!="":
        operators2_names.focus_set()
        operator1_names_lab.config(bg = "green2")
        operators_names1  = (operators_names.get()).upper()
        print(operators_names1)

    else:
        operator1_names_lab.config(bg = "orange1")
        
def operators_namesfunc(event):
    global operators_names1
    if str(operators_names.get())!="":
        operators_names1  = (operators_names.get()).upper()
        operator1_names_lab.config(bg = "green2")
        PASSWORD_OP.focus_set()
    
        print(operators_names1)
    else:
        operator1_names_lab.config(bg = "orange1")

def operators_namesdel(event):
    global operators_names1
    operators_names1  = ""
    print("operators_names1 is deleted!")

## Adding combobox FILM A REFAIR
n4 = tk.StringVar() 
 
operators_names = ttk.Combobox(starting, width = 15, textvariable = n4,font= ("Courier", 12, "bold"), state="readonly" )

operators_names['values'] =P.load_element("LISTE_operateurs") #load_list("LISTE_operateurs.txt")
  
operators_names.grid(column = 1, row = 0,padx = 10, pady=5) 
operators_names.current()
operators_names.bind("<<ComboboxSelected>>", operators_namesfunc)
operators_names.bind("<BackSpace>", operators_namesdel)
operators_names.focus_set()
"""
operators_names = Entry(starting, width = 18,relief ="sunken", font =("Helvetica",16), bg ="white")
operators_names.grid(column = 1, row = 0,padx = 10, pady=5)
operators_names.bind("<Return>", focus_func)
"""
# combobox6

PASSWORD_OP_lab = Label(starting, text= "PASSWORD:",font =("Helvetica",10,"bold"), bg =color)
#PASSWORD_OP_lab.grid(row = 1 , column= 0, padx = 10,sticky="W", pady=5)

def password_func(event):
    operators2_names.focus_set()
    

PASSWORD_OP = Entry(starting, width = 18,relief ="sunken", font =("Helvetica",12), bg ="white",show = '*')
#PASSWORD_OP.grid(row = 1 , column= 1,padx = 10, pady = 1,columnspan=2,sticky="W")
PASSWORD_OP.bind('<Return>', password_func)


#OPERATOR2 LABEL

operator2_names_lab = Label(starting, text= "OPERATEUR 02:",font =("Helvetica",10,"bold"), bg =color)
operator2_names_lab.grid(row = 2 , column= 0, padx = 10,sticky="W", pady=5)


def focus_func1(event):
    global operators2_names1
    if str(operators2_names.get())!="":
        POST.focus_set()
        operators2_names1  = (operators2_names.get()).upper()
        operator2_names_lab.config(bg = "green2")
        print(operators2_names1)
    else:
        operator2_names_lab.config(bg = "orange1")
        

# combobox5
def operators2_namesfunc(event):
    global operators2_names1
    if str(operators2_names.get())!="":
         operators2_names1  = (operators2_names.get()).upper()
         operator2_names_lab.config(bg = "green2")
         print(operators2_names1)
    else:
        operator2_names_lab.config(bg = "orange1")

def operators2_namesdel(event):
    global operators2_names1
    operators2_names1  = ""
    print("operators2_names1 is deleted!")

## Adding combobox FILM A REFAIR
n4 = tk.StringVar() 
 
operators2_names = ttk.Combobox(starting, width = 15, textvariable = n4,font= ("Courier", 12, "bold") , state="readonly")

operators2_names['values'] =P.load_element("LISTE_operateurs") #load_list("LISTE_operateurs.txt")
operators2_names.grid(column = 1, row = 2,padx = 10, pady=5) 
operators2_names.current()

operators2_names.bind("<<ComboboxSelected>>", operators2_namesfunc)
operators2_names.bind("<BackSpace>", operators2_namesdel)
"""
operators2_names = Entry(starting, width = 18,relief ="sunken", font =("Helvetica",16), bg ="white")
operators2_names.grid(column = 1, row = 1,padx = 10, pady=5)
operators2_names.bind("<Return>", focus_func1)

"""


# POST LABEL

POST_lab = Label(starting, text= "POSTE:",font =("Helvetica",10,"bold"), bg =color)
POST_lab.grid(row = 3 , column= 0, padx = 10,sticky="W", pady=5)

list_des_post=P.load_element("LISTE_DES_POST")#load_list("LISTE_DES_POST.txt")
def POSTfunc(event):
    global POST1,EQUIPE1,iqi_check
    timing = datetime.now()
    if str(POST.get())!="":

        timing = datetime.now()
        if int(list_des_post[0])<=timing.hour<int(list_des_post[1]) :
            POST1="1 er"
            iqi_check = int(list_des_post[0]) + int(list_des_post[6])
        if int(list_des_post[2])<=timing.hour<int(list_des_post[3]) :
            POST1="2 eme"
            iqi_check = int(list_des_post[2]) + int(list_des_post[6])
        if int(list_des_post[4])<=timing.hour<int(24) or int(0)<=timing.hour <int(list_des_post[5]):
            if int(list_des_post[4])<=timing.hour<int(24):
                iqi_check = int(list_des_post[4]) + int(list_des_post[6])
            else:
                iqi_check =1
                
            POST1="3 eme"
        
        EQUIPE1= POST.get()[0]
        POST_lab.config(bg = "green2")
        
        print("post=",POST1,"equipe=",EQUIPE1)
    else:
        POST_lab.config(bg = "orange1")

def POSTdel(event):
    global POST1
    POST1  = ""
    EQUIPE1= ""
    print("POST1 is deleted!")
    
n5 = tk.StringVar() 
 
POST = ttk.Combobox(starting, width = 8, textvariable = n5,font= ("Courier", 12, "bold"), state="readonly" )

POST['values'] =('A','B','C','D')
  
POST.grid(column = 1, row = 3,padx = 10, pady=5,sticky="W") 
POST.current()
POST.bind("<<ComboboxSelected>>", POSTfunc)
POST.bind("<BackSpace>", POSTdel)


# EQUIPE

# EQUIPE LABEL

PROJECT_lab = Label(starting, text= "PROJET:",font =("Helvetica",10,"bold"), bg =color)
PROJECT_lab.grid(row = 4 , column= 0, padx = 10,sticky="W", pady=5)


def PROJECTfunc(event):
    global PROJECT
    if str(Projet.get())!="":
        PROJECT  = Projet.get()
        PROJECT_lab.config(bg = "green2")
    else:
        PROJECT_lab.config(bg = "orange1")
        
    print(PROJECT)

def PROJECTdel(event):
    global PROJECT
    PROJECT  = ""
    print("PROJECT is deleted!")

    


n5 = tk.StringVar() 
 
Projet = ttk.Combobox(starting, width = 8, textvariable = n5,font= ("Courier", 12, "bold") , state="readonly")
 
Projet['values'] =P.load_element("LISTE_Projets")#load_json("projet.json")['projet']

#print("aaaa=====",a)

Projet.grid(column = 1, row = 4,padx = 10, pady=5,sticky="W") 
Projet.current()
Projet.bind("<<ComboboxSelected>>", PROJECTfunc)
Projet.bind("<BackSpace>", PROJECTdel)

##def my_popup(event):
##    path_menu.tk_popup(event.x_root,event.y_root)
##    pass
##path_menu =Menu(starting,tearoff=False)
##path_menu.add_command(label="Past",command= lambda: pathE.event_generate('<<Paste>>'))
##
##
##
##pathE = Entry(starting, width = 19,relief ="groove", font =("Helvetica",15),bg="white")
###pathE.grid(row = 4 , column= 1, pady = 5,columnspan = 3)
###pathE.insert(0,"Z:\GZ-2")
##pathE.insert(0,r"{}".format(covert_to_excel(load_list("PATH_PROJET.txt"))))
##pathE.bind("<Button-3>",my_popup)
##
###pathE.insert(0,"C:\\Users\\111\\Desktop\\EGPDF")
##
##pathE_lab = Label(starting, text= "PATH:",font =("Helvetica",10,"bold"), bg =color)
###pathE_lab.grid(row = 4 , column= 0, padx = 10,sticky="W")

var7 = IntVar()



fr1_btn= tk.Button(starting,text="ENTER",bg = btncolor,command=lambda:show_frame(PIPE_FRAME_FRAME),font =("Helvetica",10,"bold"),height = 1, width = 10)
fr1_btn.grid(row =6,column=1, padx = 15,pady = 20,sticky ="E")

################################################################################################ END STARTING FRAME ##########################################################################################


x_space=10

############################################################################################### PIPE_FRAME ############################################################################################################



# PIPE_FRAME WEDGITS

PIPE_NAME_lab = Label(PIPE_ET_DESIGNATION, text= "PIPE: ",font =("Helvetica",15,"bold"), bg =color)
PIPE_NAME_lab.grid(row = 0 , column= 0, padx =0,sticky="W")


PIPE_NAME = Entry(PIPE_ET_DESIGNATION, width = 6,relief ="sunken", font =("Helvetica",20), bg ="white")
PIPE_NAME.grid(row = 0 , column= 1, padx = x_space, pady = 10)
PIPE_NAME.focus_set()
PIPE_NAME.bind("<Return>", open_and_create_folder)
PIPE_NAME.bind("<Button-1>", open_and_create_folder)


# defect name label

#defect_name_lab = Label(PIPE_ET_DESIGNATION, text= "DESIGNATION:",font =("Helvetica",10,"bold"), bg =color)
#defect_name_lab.grid(row = 1 , column= 0, padx = 10,sticky="W")



# combobox1
font1= ("Courier", 20, "bold")

def defect_namefunc(event):
    global defect_name1
    defect_name1 = defect_name.get()
    print(defect_name1)

def defect_namedel(event):
    global defect_name1
    defect_name1 = ""
    print("defect_name1 is delated!")


n = tk.StringVar() 
defect_name = ttk.Combobox(PIPE_ET_DESIGNATION, width = 2, textvariable = n,font= font1) 
  
# Adding combobox DEFECT NAME:
defect_name['values'] =  ('D',
                          'F',
                          'K',
                          'E', 
                          'Y', 
                          'EY', 
                          'U', 
                          'S',
                          'DS',
                          'FS',
                          'D1S',
                          'F1S')
  
defect_name.grid(row = 0,column = 2, padx = x_space, pady = 10) 
defect_name.current()
defect_name.bind("<<ComboboxSelected>>", defect_namefunc)
defect_name.bind("<BackSpace>", defect_namedel)


# combobox2.5

def defect_numberfunc(event):
    global defect_number1
    defect_number1= defect_number.get()
    print(defect_number1)
    
def defect_numberdel(event):
    global defect_number1
    defect_number1= ""
    print("defect_number1 is delated!")


    
    
n1 = tk.StringVar() 
defect_number = ttk.Combobox(PIPE_ET_DESIGNATION, width = 2, textvariable = n1,font= font1) 

#defect_name,defect_number,defect_letter,defect_FAR


    
# Adding combobox NUMBER OF DEFECTS:
defect_number['values'] =('1',  
                          '2', 
                          '3', 
                          '4', 
                          '5', 
                          '6', 
                          '7',
                          '8', 
                          '9', 
                          '10',
                          '11',
                          '12',
                          '13',
                          '14',
                          '15')
  
defect_number.grid(row = 0,column = 3, padx = x_space, pady = 10) 
defect_number.current()
defect_number.bind("<<ComboboxSelected>>", defect_numberfunc)
defect_number.bind("<BackSpace>", defect_numberdel)



  

# combobox3

def defect_letterfunc(event):
    global defect_letter1
    defect_letter1 = defect_letter.get()
    print(defect_letter1)
    
def defect_letterdel(event):
    global defect_letter1
    defect_letter1 = ""
    print("defect_letter1 is deleted!")



    
    
# Adding combobox LETTER OF DEFECT
n2 = tk.StringVar() 
 
defect_letter = ttk.Combobox(PIPE_ET_DESIGNATION, width = 1, textvariable = n2,font= font1 )

defect_letter['values'] =('A',  
                          'B', 
                          'C', 
                          'D', 
                          'E', 
                          'F', 
                          'G',
                          'H', 
                          'I', 
                          'J',
                          'K',
                          'L',
                          'M'
                          )
  
defect_letter.grid(row = 0,column = 4, padx = x_space, pady = 10) 
defect_letter.current()
defect_letter.bind("<<ComboboxSelected>>", defect_letterfunc)
defect_letter.bind("<BackSpace>", defect_letterdel)



# combobox4
def defect_FARfunc(event):
    global defect_FAR1
    defect_FAR1  = defect_FAR.get()
    print(defect_FAR1)

def defect_FARdel(event):
    global defect_FAR1
    defect_FAR1  = ""
    print("defect_FAR1 is deleted!")

    
# combobox2

def SOUNDAGEfunc(event):
    global SOUNDAGE1
    SOUNDAGE1= SOUNDAGE.get()
    print(SOUNDAGE1)
    
def SOUNDAGEdel(event):
    global SOUNDAGE1
    SOUNDAGE1= ""
    print("defect_number1 is delated!")


    
    
n9 = tk.StringVar() 
SOUNDAGE = ttk.Combobox(PIPE_ET_DESIGNATION, width = 3, textvariable = n9,font= font1) 

#defect_name,defect_number,defect_letter,defect_FAR


    
# Adding combobox NUMBER OF DEFECTS:
SOUNDAGE['values'] =('SD', 
                     'SD1', 
                     'SD2', 
                     'SD3', 
                     'SG',
                     'SG1', 
                     'SG2',
                     'SG3')
  
SOUNDAGE.grid(row = 1,column = 1, padx = x_space, pady = 10) 
SOUNDAGE.current()
SOUNDAGE.bind("<<ComboboxSelected>>", SOUNDAGEfunc)
SOUNDAGE.bind("<BackSpace>", SOUNDAGEdel)


## Adding combobox FILM A REFAIR
n3 = tk.StringVar() 
 
defect_FAR = ttk.Combobox(PIPE_ET_DESIGNATION, width = 3, textvariable = n3,font= ("Courier", 20))

defect_FAR['values'] =('.',  
                       '..', 
                       '...', 
                       '....' )
  
defect_FAR.grid(row =1,column = 2, padx = x_space, pady = 10,columnspan=2,sticky="w") 
defect_FAR.current()
defect_FAR.bind("<<ComboboxSelected>>", defect_FARfunc)
defect_FAR.bind("<BackSpace>", defect_FARdel)



#check

    

var2 = IntVar()
def check(event):
    if var2.get()==1:
        RM1.config(bg =color)
    else :
        RM1.config(bg="red")
 

RM1= Checkbutton(PIPE_ET_DESIGNATION, text = "RM1", variable = var2, bg =color,font= font1)
RM1.grid(row = 1 , column= 3,padx = x_space, pady = 10,columnspan=2)
#RM1.select()

RM1.bind('<Return>',check)
RM1.bind('<Button-1>',check)

l_variable = Spinbox(PIPE_ET_DESIGNATION,from_=0, to = 5 ,bg ="white",increment =.1,width = 4, font =("Helvetica",15),buttonbackground = "orange" ,relief ="sunken", highlightcolor= "yellow")
#l_variable.grid(row = 2 , column= 2, padx = 5,sticky="W",pady = 10)
l_variable.delete(0,"end")
l_variable.insert(0,1)
# INTEGRATION BUTTON
#######################################################################################################################################################################
###################################################################### MAIN FRAME #######################################################################################

# line 2

INTGRATIONe_lab = Label(PIPE_FRAME_FRAME, text= "INT/TUBE:",font =("Helvetica",10,"bold"), bg =color)
INTGRATIONe_lab.grid(row = 1 , column= 0, padx = 10, pady = 10,sticky="W")


DEFECT_NUMB_lab = Label(PIPE_FRAME_FRAME, text= i,font =("Helvetica",10,"bold"), bg =color)
DEFECT_NUMB_lab.grid(row = 1 , column= 1, padx = 10, pady = 10)

###################################################################### ACTIONS_FRAME: #######################################################################################


#actions_lab = Label(ACTIONS_FRAME, text= "ACTIONS:",font =("Helvetica",10,"bold"), bg =color)
#actions_lab.grid(row = 4 , column= 0, padx = 10, pady = 10,sticky="W")


# buttons
INTEGRATION_btn = Button(PIPE_FRAME_FRAME, text = "INTÉGRE", bg =btncolor,activebackground="YELLOW1",font =("Helvetica",10,"bold"),height = 2, width = 13)
INTEGRATION_btn.grid(row = 2 , column= 0,padx = 10, pady = 10,columnspan=1)
INTEGRATION_btn.bind('<Return>',integration)
INTEGRATION_btn.bind('<Button-1>',integration)

finish_tube = Button(PIPE_FRAME_FRAME, text = "INSÈRE", bg =btncolor,activebackground="orange",font =("Helvetica",10,"bold"),height = 2, width = 13)
finish_tube.grid(row = 2 , column= 1,padx = 5, pady = 10)
finish_tube.bind('<Return>',tube_finished)
finish_tube.bind('<Button-1>',tube_finished)

report_closed = Button(PIPE_FRAME_FRAME, text = "CLÔTURE", bg =btncolor,activebackground="red",font =("Helvetica",10,"bold"),height = 2, width = 13)
report_closed.grid(row = 2 , column= 2,padx = 5, pady = 10)
report_closed.bind('<Return>',report_closed_func)
report_closed.bind('<Button-1>',report_closed_func)

#QUIT_btn = Button(ACTIONS_FRAME, text = "QUIT", bg =btncolor,activebackground="RED",font =("Helvetica",10,"bold"),height = 2, width = 13, command = quitt)
#QUIT_btn.grid(row = 5 , column= 2,padx = 5, pady = 10)
#QUIT_btn.bind('<Return>',quitt)
#QUIT_btn.bind('<Button-1>',quitt)

var5 = IntVar()

##use_saerch_image = Checkbutton(PIPE_FRAME_FRAME, text = "AUTO", variable = var5, bg =color)
##use_saerch_image.grid(row = 1, column= 2,padx = 30, pady =10,sticky="w")
##use_saerch_image.select()

continueing_checkbtn= Checkbutton(PIPE_FRAME_FRAME, text = "R_INCOMPLET",font =("Helvetica",8,"bold"), variable = var7, bg =color)
continueing_checkbtn.grid(row = 6 , column= 0,padx = 10, pady = 5,sticky="W")
continueing_checkbtn.bind('<Button-1>',check5)


j_variable_lab = Label(PIPE_FRAME_FRAME, text= "Line N°:",font =("Helvetica",10,"bold"), bg =color)
j_variable = Spinbox(PIPE_FRAME_FRAME,from_=1, to = max_line ,bg ="white",increment =1,width =5, font =("Helvetica",10),buttonbackground = "orange" ,relief ="sunken", highlightcolor= "yellow")
#j_variable.grid(row = 4 , column= 1, pady = 10,columnspan = 2)
#j_variable.delete(0,"end")
#j_variable.insert(0,"8")
#j_variable.bind("<<SpinboxSelected>>",j_variablefunc)

k_variable_lab = Label(PIPE_FRAME_FRAME, text= "Report N°:",font =("Helvetica",10,"bold"), bg =color)
k_variable = Spinbox(PIPE_FRAME_FRAME,from_=1, to = max_line ,bg ="white",increment =1,width =5, font =("Helvetica",10),buttonbackground = "orange" ,relief ="sunken", highlightcolor= "yellow")



j_variable_btn= tk.Button(PIPE_FRAME_FRAME,text="VALIDE",bg = btncolor,command=j_variablefunc,font =("Helvetica",10,"bold"),height = 1, width = 10)


# line 3
devlabel = Label(PIPE_FRAME_FRAME, text= "DEVELOPED BY BOUZID YASSINE CND-RT-II 2020",font =("Helvetica",10), bg =color)
devlabel.grid(row = 12 , column= 0, columnspan = 7, pady = 15, padx= 60)

fr3_btn= tk.Button(PIPE_FRAME_FRAME,text="RETOUR",bg = btncolor,command=lambda:show_frame(starting_FRAME),font =("Helvetica",10,"bold"),height = 1, width = 10)
fr3_btn.grid(row = 11 , column= 1,padx = 10,pady=10)


############################################################################################### END PIPE_FRAME ############################################################################################################




icon = """AAABAAEA4eEAAAEACACs6AAAFgAAACgAAADhAAAAwgEAAAEACAAAAAAAZMgAAAAAAAAAAAAAAAEAAAABAAA4/f8AGxYZAP///wDu7u4A7e3tADn//wAAAAAA+fn5APT09AD7+/sAFwAAADCvsQAVAAAAGAAAABsVGAATAAAAJ4mKAC+VlgAaERQAGQoOABYQFAAbAgkAztjZAAoAAAAbAAAAycjJABkRFAAfbW4AABcZABkGCwAPAAAAGAwPADjx8gAfAAAANcnJADPW1wAWFxkAI0ZIADfw8QDY3t8Ay9DQAGBdXQBNSUoAPDY3AB0kJgBrb3AAABEVANrZ2QB9enoAH3R2ABpMTQBUWlsAuLe3ADJydAAyvL4AK01PADk9PgAjW1wAJDEzADF/gQCPlJUAI2VnAJmXlwA35+gAIxEWADWztABqZ2gAp6usAGt8fgClo6MAN0lLADxGSAAmHR4Asby9AJCVlgCRoaIAJywtAKCvsQAbMjQALcDBADpPUQAyeXoAWGttAC2iowAtZWYAGz5AAIWFhgAAHyIAc46QABYqKwA7q6wAFh4gADaTlAAuPD4ALi0vADFeYAAwVFYAM2BiAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICBAMDAwMDAwMEBwICAgICAgICCQQDAwgHBwkJAgICBAQnFhYWKCgWFhYWBAQEAgICAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICBAMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICBAMDAwMDAwMEBwICAgICAgICCQkIAwQ0NFhEQikpKytMHBwcBgYuHBwcKysrKSktREM0NAMIBwICCAgIBAQEAwMIAgICAgICAgICBAMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICBAMDAwMDAwMEBwICAgICBElJRTBXBgYGBgYGFxQUFBQUEhISDg4OEhIaFBQUFBcXBgYGBgYGVzBFSUkIAgIJAwMIAgICAgICAgICBAMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICBAMDAwMEAwgHAgQnNFY4XgYGBhcSDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEaEhcGBgZIRzwvJwQCAgICAgICAgICBAMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICBAMDAwMCAgJNVjgGBgYUFAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEUHgYGBjhWGQICAgICAgICBAMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAwcJAi9FLAYGBhQOAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDhQGBgYsRSgCAgICBAMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICBwREKRQGGgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEaBhQpRAQJCQgDBAMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAi8wKgYGFAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEUBgYqRS8CCQMEAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAABAQEBAQEBAQEBwICAgICAgICCQQEBAQEBAQEBAICAgICAgICAgcEBAQEBAQEBAMCAgICAgICAgIIBAQEBAQEBAQIAgICAgRFRh8GFAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEUBh9QRQQCCAQEBwICAgICAgICCQQEBAQEBAQEBAICAgICAgICAgcEBAQEBAQEBAMCAgICAgICAgIIBAQEBAQEBAQIAgICAgICAgICAAAABwcHBwcHBwcHCAgICAgICAgICAcHBwcHBwcHBwgICAgICAgICAcHBwcHBwcHBwcICAgICAgICAgHBwcHBwcHBwcHAgIZVl4GHgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDhITFRUKDQ0MDAoNDQwMGBUVExIOAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4eBl5KAwICCAgICAgICAgICAcHBwcHBwcHBwgICAgICAgICAcHBwcHBwcHBwcICAgICAgICAgHBwcHBwcHBwcHCAgICAgICAgIAAAAAgICAgICAgICCAQDAwMDAwMECAICAgICAgICAgMEAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIIGTgGBhQBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4fHQoMDA8MEzoyGxsQEBAQEBAQEBAQEBsbMjpADwoMDAodEg4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBARQGBkQZAgcDBAQDAwMECAICAgICAgICAgMEAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICBxk4Bh4BAQEBAQEBAQEBAQEBAQEBAQEBAQEOGh0NDB9ZJSU1Cz8/Pz8gAAAFBQUFBQUFBQUFBQAAID8/IyILNSUlWR8MDR0aAQEBAQEBAQEBAQEBAQEBAQEBAQEBDgYGMy8JCQMEAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgJDTgYeAQEBAQEBAQEBAQEBAQEBAQEBAQ4TDQ8PQDlTQSIABQUFBQUFBQUFBQUAAAAAAAAAAAAAAAUFBQUFBQUFBQUFACJBETlADA8NHw4BAQEBAQEBAQEBAQEBAQEBAQEBFwYpGQIJBAMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgIISToGHwEBAQEBAQEBAQEBAQEBAQEBAQEfDRQyPUEmIAAFBQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUFBQAgJhE9VRQKHwEBAQEBAQEBAQEBAQEBAQEBAQEXBioZCQcECAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQCAhlHBhQBAQEBAQEBAQEBAQEBAQEBAQ4NDwwsCyMgBQUFBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFBQUFPyMLLB4eDQ4BAQEBAQEBAQEBAQEBAQEBAR4GQgQCCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMECAIWOAYeAQEBAQEBAQEBAQEBAQEBAQ4dDyExIgAFBQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUFBQVaMSEPDBIBAQEBAQEBAQEBAQEBAQEBFwYtJwICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDBAMCCUQsFwEBAQEBAQEBAQEBAQEBAQETDRQ3Nj8FBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBT82NxQKEgEBAQEBAQEBAQEBAQEBAQEGLD4CAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAQDAwMDAwMECAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMEBwJDDgYUAQEBAQEBAQEBAQEBAQEaDw9VCyMFBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFBSMLIQ8NDgEBAQEBAQEBAQEBAQEBGgYqSQICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAACQkJCQkJCQkJCAgICAgICAgICAkJCQkJCQkJCQgICAgICAgICAcJCQkJCQkJCQkICAgCBEIGFAEBAQEBAQEBAQEBAQEBGg8SUU8ABQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFADY5DAoOAQEBAQEBAQEBAQEBAQEXBkIEAgkJCQgICAgICAgICAcJCQkJCQkJCQkICAgICAgICAgHCQkJCQkJCQkHCAgICAgICAgIAAAABAQEBAQEBAQEBwICAgICAgICCQQEBAQEBAQEBAICAgICAgICAgcEBAQEBAQEBAMCAgJLXgYBAQEBAQEBAQEBAQEBAR8PDREjBQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBQUFBQUFBQUFBQUFBQUFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUFIlQeDxoBAQEBAQEBAQEBAQEBDgZdNAIIBAICAgICAgICAgcEBAQEBAQEBAMCAgICAgICAgIIBAQEBAQEBAQIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDBAMCLzMGFAEBAQEBAQEBAQEBAQETDFURAAUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBQUFBQUFBQUAJj8jIiIiIiIiIiIjIz8mAAUFBQUFBQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQU/EFsMEgEBAQEBAQEBAQEBAQEUBkQECQICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMECAlJXgYBAQEBAQEBAQEBAQEBEgxOECAFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBQUFACAgP1M7PT05Mk4SEwwMGA0NDRgOEk4yOT09XAs/ICAABQUFBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBT9UWwwOAQEBAQEBAQEBAQEBAQZeKAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwQHCC0GFAEBAQEBAQEBAQEBARoMDBEFBQUAAAAAAAAAAAAAAAAAAAAAAAAABQUFBQUgIwsxXQ0KDw8PDRUdExIODgEBDg4ODg4SDhITHRUNDw8MDQ0lUwsjIAUFBQUFAAAAAAAAAAAAAAAAAAAAAAAAAAUFJjsYDA4BAQEBAQEBAQEBAQEUBkUJAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwIvWQYBAQEBAQEBAQEBAQEBCh0QIAUFAAAAAAAAAAAAAAAAAAAAAAAAAAUFAD8/OyVZFA0NHw4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4TDRgkWSU7PyYABQAAAAAAAAAAAAAAAAAAAAAAAAAABQUmNQwTAQEBAQEBAQEBAQEBDgZQCAICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAksPHgEBAQEBAQEBAQEBARMeMiIFBQAAAAAAAAAAAAAAAAAAAAAAAAUFBSBTG0gMDBMSDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEOEhMMDFUbUyAFBQUAAAAAAAAAAAAAAAAAAAAAAAUFBTZbDxIBAQEBAQEBAQEBAQEXXhkCAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEBAgCRAYUAQEBAQEBAQEBAQEOHk5BBQUAAAAAAAAAAAAAAAAAAAAABQUFBSMRQB4PFRoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR8VDx5AESMFBQUFAAAAAAAAAAAAAAAAAAAAAAUFXFsMDgEBAQEBAQEBAQEBHwZYAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcECQRCBg4BAQEBAQEBAQEBAQoNED8FAAAAAAAAAAAAAAAAAAAAAAUFBQARPSQKHwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEfCiQ9EQAFBQAAAAAAAAAAAAAAAAAAAAAABT89DB8BAQEBAQEBAQEBAQ4GPAICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAABAQEBAQEBAQEBwICAgICAgICCQQEBAQEBAQEBAICAgICAgICAggCJysGAQEBAQEBAQEBAQEaD1U2BQUAAAAAAAAAAAAAAAAAAAAFBQVPMSEPFRoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARoVDyExTwUFBQAAAAAAAAAAAAAAAAAAAAUFCyQKDgEBAQEBAQEBAQEOBkIDAggEBAQEBAQEBAMCAgICAgICAgIIBAQEBAQEBAQIAgICAgICAgICAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDAwIoWR4BAQEBAQEBAQEBAR8PNQUFAAAAAAAAAAAAAAAAAAAAAAUgCzIPDRoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARoNDFVTJgUFAAAAAAAAAAAAAAAAAAAABT85DxoBAQEBAQEBAQEBAQYqJwICAgICAgICAgIDAwMDAwMDAwMJAgICAgICAgIHAwMDAwMDAwMDAAAAAgICAgICAgICCAQDAwMDAwMDCAICAgICAgICAgMDAwMDAwMEAigGHwEBAQEBAQEBAQEBDyEiBQUAAAAAAAAAAAAAAAAABQUFBVpVDAoOAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDhMMThEmBQUAAAAAAAAAAAAAAAAAAAUFUwwdAQEBAQEBAQEBAQEXVycCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwQCNAYaAQEBAQEBAQEBAQ4MPSAFBQAAAAAAAAAAAAAAAAAFBSNTCg8TAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBEg8KECMFBQAAAAAAAAAAAAAAAAAFBT8lDwEBAQEBAQEBAQEBHwYoAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDCAJWBgEBAQEBAQEBAQEBEx4RAAUAAAAAAAAAAAAAAAAABQUgUyQKGgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEaCiRTIAUFAAAAAAAAAAAAAAAAAAU/Gw0SAQEBAQEBAQEBARQGNAICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAj4GAQEBAQEBAQEBAQEKJE8FBQAAAAAAAAAAAAAAAAAFBVMVDxoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARoPFVMABQAAAAAAAAAAAAAAAAAFBVMeEwEBAQEBAQEBAQEaBhkCAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwgHVgYBAQEBAQEBAQEBAQ8lJgUAAAAAAAAAAAAAAAAAAAAFIw0fAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHRMgBQAAAAAAAAAAAAAAAAAAAAUjWx0BAQEBAQEBAQEBDgY0AgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDBAI+BhQBAQEBAQEBAQEBHkYgBQAAAAAAAAAAAAAAAAAAAAAFBTsMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDVoFAAAAAAAAAAAAAAAAAAAAAAUFPyQMAQEBAQEBAQEBARQGGQICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMEAlYGAQEBAQEBAQEBARIPEQUFAAAAAAAAAAAAAAAAAAAAAAAABSYkHwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKOQUAAAAAAAAAAAAAAAAAAAAAAAAABQBfDw4BAQEBAQEBAQEOBjQCAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAQEBAQEBAQECAICAgICAgICAgMCVgYBAQEBAQEBAQEBGgwRBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAUQDAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4KIwUAAAAAAAAAAAAAAAAAAAAAAAAAAAUgGw8OAQEBAQEBAQEBGgYoAgIEBAQEBAQEBAQJAgICAgICAgIHBAQEBAQEBAQEAAAABwcHBwcHBwcHBwcHBwcHBwcHBwcHBwcHBwcHCAIZBgEBAQEBAQEBAQESDDYFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAUgVRMBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQwbBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAFABEPDgEBAQEBAQEBARQGJwIIBwcHBwcHBwcHBwcHBwcHBwcHBwcHBwcHBwcHAAAABAQEBAQEBAQEBwICAgICAgICCQQEBAQEBAQEAhkGDgEBAQEBAQEBAR8fNgUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFNg8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDhgmBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQVTDw4BAQEBAQEBAQEfVycCAgICAgICAgIIBAQEBAQEBAQIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwQJJy4aAQEBAQEBAQEBDg82BQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSUNAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDxEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFEA8BAQEBAQEBAQEBFyoDAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwgEKxcBAQEBAQEBAQESDyMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSMkEgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEdMiAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBU8PDgEBAQEBAQEBAQZCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwJCBgEBAQEBAQEBARISNgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUxDAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEPCwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUQDw4BAQEBAQEBAQEGPAICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAi0GAQEBAQEBAQEBGgw2BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU/IRoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ1RBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFUw8OAQEBAQEBAQEBBkMCAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwQCSwYBAQEBAQEBAQESDDYFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFCw8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDhUiBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBREKDgEBAQEBAQEBFAYZAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDBAcvDw4BAQEBAQEBAQEPEQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFADoTAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBChsFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQAbDwEBAQEBAQEBARdeCAICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDCAcUFAEBAQEBAQEBDgwRBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSIMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQESWSYFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUgXwwBAQEBAQEBAQEXUAkCAgIIAwMDAwMDAwQIAgICAgICAgICAAAAAwMDAwMDAwMDBwICAgICAgICCQMDBy0GAQEBAQEBAQEBDz0FBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSA9DAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEPWgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFACQdAQEBAQEBAQEBBkUCAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAgICAgICAgICCAQEBAQEBAQECAICNAYOAQEBAQEBAQEKJSAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUgDw4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR1OBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABT9bEwEBAQEBAQEBDgYoCQQJAgICAgICAgIHBAQEBAQEBAQEAAAAAgICAgICAgICCAMDAwMDAwMDCAIoLhoBAQEBAQEBARMkJgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFMQoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDgwiBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUjHhIBAQEBAQEBARReBAgJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAJIHgEBAQEBAQEBDh5PBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFP1kSAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBCjkFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFUw0BAQEBAQEBAQEGRAIJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMEAkQGAQEBAQEBAQEBDxEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABUEeAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEaGD8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABRsPAQEBAQEBAQEBBjQCAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwQHLwYUAQEBAQEBAQEPPQAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQA6EwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEPEQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABT8lHQEBAQEBAQEBFF0EAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMIXR4BAQEBAQEBARMhIAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVPDw4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR8kIAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU/DBoBAQEBAQEBAQZCAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDBAJDBgEBAQEBAQEBDhQiBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFPQwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQxTBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFUw8OAQEBAQEBAQ4GSQICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDCAgSFAEBAQEBAQEBDDUFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIyEOAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBEzIgBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBTkKAQEBAQEBAQEeKgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAQEBAQEAlIGAQEBAQEBAQEdVQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABVMPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEOHiMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABT8kHwEBAQEBAQEBBj4CAgICAgIHBAQEBAQEBAQEAAAACAgICAgICAgIBwkJCQkCKAYBAQEBAQEBAQ4YNgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUSHwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQENPQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAULDA4BAQEBAQEBGiwnCAgICAgICQkJCQkJCQkJAAAAAwMDAwMDAwMEBwICAgICKx4BAQEBAQEBAQ8QBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU2DAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR9ZPwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFPQwBAQEBAQEBAQYtAgQDAwQIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgJNBgEBAQEBAQEBEyU/BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUAMh0BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ8RBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFP1sSAQEBAQEBAQEGBAcDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAghHGgEBAQEBAQEBHjYFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIxcBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBEiEmBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABVwPAQEBAQEBAQEXQgkEAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAkUGAQEBAQEBAQEPUQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBRAKAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBD08FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVbEwEBAQEBAQEBBhkJBAMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAl4UAQEBAQEBARoSIwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSMMGgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKJQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU2DA4BAQEBAQEBHioCAwQIAgICAgICAgICAAAAAwMDAwMDAwMEBwICPAYBAQEBAQEBAQ8QBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVTHgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMIwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUAVAoBAQEBAQEBAQYZCQQIAgICAgICAgICAAAAAwMDAwMDAwMEBwIIBhoBAQEBAQEBH1kgBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUmThMBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ8RBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFJhgOAQEBAQEBARcpCQMIAgICAgICAgICAAAAAwMDAwMDAwMEBwJDBgEBAQEBAQEBDFMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFQR4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBGh4ABQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBTsMAQEBAQEBAQEGLwcIAgICAgICAgICAAAACAgICAgICAgIBwg4FAEBAQEBAQETWQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABVUNAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDFMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSZbEgEBAQEBAQEXMwIIBwcHBwcHBwcHAAAAAgICAgICAgICAjQGAQEBAQEBAQEPCwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSISDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQETJSAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVUDAEBAQEBAQEBBhkCBAQEBAQEBAQEAAAAAgICAgICAgICAgYfAQEBAQEBARMlIAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU1CgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4eIgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU/WxIBAQEBAQEBBkQCBAMDAwMDAwMDAAAAAgICAgICAgICGQYBAQEBAQEBAQ9TBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUmDBIBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ0bBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFEA8BAQEBAQEBDgYDCAMDAwMDAwMDAAAAAgICAgICAgICVgYBAQEBAQEBGg0FBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFEQwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBEh8jBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFPx4OAQEBAQEBAQZKAgQDAwMDAwMDAAAAAgICAgICAgIvBgEBAQEBAQEBChEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBVUTAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDzEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABVQKAQEBAQEBARReBAMDAwMDAwMDAAAAAgICAgICAgJFBgEBAQEBAQESTj8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQseAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQETVT8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSIMDgEBAQEBAQEGRQkEAwMDAwMDAAAAAgICAgICAghMGgEBAQEBAQEPUQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQA5CgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEeCwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU5DQEBAQEBAQEeUAIDAwMDAwMDAAAAAgICAgICAkkGDgEBAQEBAQEMIwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUjDw4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARNOBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU2DwEBAQEBAQEOHy8IAwMDAwMDAAAAAgICAgICAjAGAQEBAQEBARU9AAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFGwoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQw2BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUgIRIBAQEBAQEBBkUJBAQEBAQEAAAABAQEBAQDCC4SAQEBAQEBAQxBBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFPywfAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDxAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFCwoBAQEBAQEBFCoJAgICAgICAAAAAwMDAwMHNAYBAQEBAQEBEls/BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQsPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEODz8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIxQaAQEBAQEBAQYEAgICAgICAAAAAwMDAwQCLRcBAQEBAQEBDBsFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQBZEwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMEAUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABTcMAQEBAQEBAQZEAgICAgICAAAAAwMDAwMEKxQBAQEBAQEBDCMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUiCg4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARM6IAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABTYMAQEBAQEBARQpAgICAgICAAAAAwMDBAk0BgEBAQEBAQESTj8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFNwoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQw2BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSMMAQEBAQEBAQEUKAICAgICAAAAAwMDBAIzHgEBAQEBAQEPEQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFPx8OAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHTcgBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUxDQEBAQEBAQEGRQICAgICAAAAAwMDAwIGGgEBAQEBAQEeIgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBREPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDz8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVaHgEBAQEBAQEfWwICAgICAAAAAwMDCAQGAQEBAQEBARoKBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUKGgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKPQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFHgEBAQEBAQEBBgICAgICAAAABAQECUoGAQEBAQEBAQ0QBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVTDQEBAQEBAQEBAQEBARoaGhoaGg4BAQEBAQEBAQEBAR8sIwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFLB8BAQEBAQEBBhkCAgICAAAABwcJBykUAQEBAQEBAQwiBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUgJQoBAQEBAQESDQ8PDxghISEhIQwPHg8dGgEBAQEBAQ8RBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFCwoBAQEBAQEBBlYCCAgIAAAAAgICJy4aAQEBAQEBEiwjBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIwwOAQESDA1AMRE/IAUFBQUFBQUgNhE9QAwMEgEBE1UABQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIhgOAQEBAQEBFDgEAwMDAAAAAgICLwYBAQEBAQEBHTkFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBTENEg9bGyM/BQUFBQAAAAAAAAAFBQUAPzY3FAoaDEEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFP04SAQEBAQEBDgYnAwMDAAAAAgICRAYBAQEBAQEBDxEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSMeIVM/BQUFAAAAAAAAAAAAAAAAAAAFBQUFIxEPNwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABT0NAQEBAQEBAQYvCQQDAAAAAgICRx4BAQEBAQEBDyMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU2PwUFAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUjPwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABREPAQEBAQEBAQY8AgQDAAAAAgIESBQBAQEBAQEaCiAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSYPDgEBAQEBAR5HAgQDAAAAAgIvDgEBAQEBAQETJQAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSBAGgEBAQEBARRICAgDAAAAAgI0BgEBAQEBAQENEAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQA5HQEBAQEBAQEGSQgDAAAAAgJFBgEBAQEBAQEKCwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAURDQEBAQEBAQEGSQgDAAAACQlCFwEBAQEBAQ4TIgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVBDAEBAQEBAQEGRQIIAAAABAhGHwEBAQEBARJOBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFBQUFBQUFBQUFBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUiHw4BAQEBAQEXMAICAAAACAgGDgEBAQEBARNOBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFBT9BPUgKDQ0NCkgbIj8FBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUAWRIBAQEBAQESVwcCAAAACAQGAQEBAQEBAR0lBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUjOU4KCh8BAQEBAR8KDFU9PwUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFJR0BAQEBAQEOBggCAAAABxkGAQEBAQEBAQw7BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFPzEMChIBAQEBAQEBAQEBAR8MITEgBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFJR0BAQEBAQEBBgMCAAAAB0MGAQEBAQEBAQxPBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQVPEw0OAQEBAQEBAQEBAQEBAQEBGg0sTwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFNQwBAQEBAQEBBjQCAAAACUQGAQEBAQEBAQ0jBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBTseHwEBAQEBAQEBAQEBAQEBAQEBAQEdFAsFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFCwwBAQEBAQEBBjQCAAAACS0GAQEBAQEBDgo/BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFUR4OAQEBAQEBAQEBAQEBAQEBAQEBAQEBEh4LBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIgoBAQEBAQEBBkMCAAAAAikXAQEBAQEBGiw/BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVTDQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARIUTwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIw8BAQEBAQEBBkQCAAAACCoUAQEBAQEBEzIgBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSMeEgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQETJD8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFP0AOAQEBAQEBFy0CAAAABCsUAQEBAQEBFRsABQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFADkVAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDTEFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFPzoSAQEBAQEBFykCAAAAJywaAQEBAQEBFRsABQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFEQ8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDgwjBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFJjITAQEBAQEBFCkCAAAAFhwOAQEBAQEBFTEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU/ThIBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQo5BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFABsVAQEBAQEBFCsEAAAAKAYOAQEBAQEBDRAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVPCgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4BIwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFABsVAQEBAQEBFCsEAAAAKAYBAQEBAQEBChAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUlHQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKUwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABRAYAQEBAQEBFCsEAAAAGQYBAQEBAQEBDxEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUPDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEfGAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABRAMAQEBAQEBGhwWAAAAGQYBAQEBAQEBDxEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABT8PAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBCgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABRAMAQEBAQEBDhwWAAAAGQYBAQEBAQEBDxEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABVoeAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHj8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABRANAQEBAQEBDhwWAAAAGQYBAQEBAQEBChEFAAAAAAAAAAUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQAAAAAAAAAABREPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDyIFAAAAAAAAAAUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQAAAAAAAAAABRANAQEBAQEBDhwWAAAAGQYBAQEBAQEBDBEFAAAAAAAAAAUjCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwtTPwUAAAAAAAAABREPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDyIFAAAAAAAAAAVPCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLCwsLPwUAAAAAAAAABRAKAQEBAQEBDhwWAAAAGQYBAQEBAQEBDxEFAAAAAAAAAAVTHg0KCg0NDQoKDQ0NCgoNDQ0KCg0NDQoKDQ0NCgoNDQ0KCg0NDQoKDQ0NCgoNDQ0KCg0NDQoKDQ0NCgoNDQ8eTwUAAAAAAAAABREPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDyIFAAAAAAAAAAUbDw0KCg0NDQoKDQ0NCgoNDQ0KCg0NDQoKDQ0NCgoNDQ0KCg0NDQoKDQ0NCgoNDQ0KCg0NDQoKDQ0NCgoNDQ8PTwUAAAAAAAAABRAMAQEBAQEBDhwWAAAAGQYBAQEBAQEBDxEFAAAAAAAAAAVTDQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKTwUAAAAAAAAABVoeAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDyIFAAAAAAAAAAUxDQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMIgUAAAAAAAAABRAMAQEBAQEBDhwWAAAAGQYBAQEBAQEBDxEFAAAAAAAAAAVPCgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKUwUAAAAAAAAABUEeAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHgAFAAAAAAAAAAVdHwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMIgUAAAAAAAAABRANAQEBAQEBDhwWAAAAGQYBAQEBAQEBDxEFAAAAAAAAAAVPCgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQENMQUAAAAAAAAABT8PAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEODwUAAAAAAAAAAAVAHwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMIgUAAAAAAAAABRANAQEBAQEBDhwWAAAAKAYBAQEBAQEBChAFAAAAAAAAAAVPDAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEfSAUAAAAAAAAAAAUNGgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEdJQUAAAAAAAAABQUNGgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMIgUAAAAAAAAABRAKAQEBAQEBFEwnAAAAKAYOAQEBAQEBDRAFAAAAAAAAAAVPDwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDAUAAAAAAAAAAAUQCgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKUwUAAAAAAAAABT8PAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4NIgUAAAAAAAAFABsVAQEBAQEBFCsEAAAAFhwOAQEBAQEBFTEFAAAAAAAAAAUiDAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHgAFAAAAAAAAAAUiIQ4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARpbIwUAAAAAAAAABUEeAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4SIgUAAAAAAAAFABsVAQEBAQEBFCsEAAAAJywaAQEBAQEBFRsABQAAAAAAAAUiDg4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDxEFAAAAAAAAAAUFOQ0BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ8QBQAAAAAAAAAABTEMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARJOPwUAAAAAAAAFIDITAQEBAQEBFykCAAAABCsUAQEBAQEBFRsABQAAAAAAAAUjEg4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBFRsABQAAAAAAAAAFIgwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDgo/BQAAAAAAAAAFPzoSAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARMyIAUAAAAAAAAFPzoSAQEBAQEBFykCAAAACCoUAQEBAQEBHT0gBQAAAAAAAAUjLBIBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDiE/BQAAAAAAAAAFBTUMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDBAFAAAAAAAAAAAFIwwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR05BQAAAAAAAAAFPxMOAQEBAQEBF0ICAAAAAikXAQEBAQEBEkw/BQAAAAAAAAUgMhMBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQwLBQAAAAAAAAAABSA6EwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKJQUFAAAAAAAAAAAFVAwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARU9BQAAAAAAAAAFPw0OAQEBAQEBBkQJAAAAAkIGAQEBAQEBDh0/BQAAAAAAAAUFOQoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARNZBQUAAAAAAAAAAAU2IRIBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR1ZIwUAAAAAAAAAAAUmWxIBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ87BQAAAAAAAAAFIwoBAQEBAQEBBlgJAAAAAkQGAQEBAQEBAQojBQAAAAAAAAAFUQ8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQENWgUAAAAAAAAAAAAFIg8fAQEBAQEBAQEBAQEBAQEBAQEBAQEBDFkjBQAAAAAAAAAABQU7DAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ8RBQAAAAAAAAAFCwwBAQEBAQEBBjQHAAAAAkMGAQEBAQEBAQxPBQAAAAAAAAAFOw8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQETOiAFAAAAAAAAAAAABTYsDAEBAQEBAQEBAQEBAQEBAQEBAQEMVSMFAAAAAAAAAAAABSYYDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ8jBQAAAAAAAAAFNQwBAQEBAQEBBjQHAAAAAhkGAQEBAQEBAQw7BQAAAAAAAAAFTw8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDxEFAAAAAAAAAAAAAAU/Mg8TAQEBAQEBAQEBAQEBAQEBDAw1IAUAAAAAAAAAAAAABTUMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDgwgBQAAAAAAAAAFJR0BAQEBAQEBBgQIAAAAAgQGAQEBAQEBAQxhBQAAAAAAAAAFIAwOAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBGg4iBQAAAAAAAAAAAAAFBQslDx8BAQEBAQEBAQEBARMMOSIFBQAAAAAAAAAAAAAFQQwOAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBGiEgBQAAAAAAAAAFJR0BAQEBAQEOBgMDAAAAAgMGDgEBAQEBARNOBQAAAAAAAAAFIBgaAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ1VBQUAAAAAAAAAAAAABQUgCywPHg0YFR0VGAoPD042BQUFAAAAAAAAAAAAAAU/IRIBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBE2AABQAAAAAAAAUAWRIBAQEBAQESVwgDAAAAAgcuGgEBAQEBARJOBQUAAAAAAAAFACUTAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMNQUFAAAAAAAAAAAAAAAFBQUFWhAxPTkbMRBPBQUFAAAAAAAAAAAAAAAABQA5DAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDREFAAAAAAAAAAUiHw4BAQEBAQEXMAkEAAAACAczHgEBAQEBARIkIwUAAAAAAAAFBRANAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDTsFBQAAAAAAAAAAAAAAAAAABQUFACAABQUFAAAAAAAAAAAAAAAAAAAFBRsMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDAsFAAAAAAAAAAVBDAEBAQEBAQEGRQIJAAAABAJFBgEBAQEBAQEKQQUAAAAAAAAABUEMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDh5TBQUAAAAAAAAAAAAAAAAAAAAABQUFAAAAAAAAAAAAAAAAAAAAAAUFOwoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQESJCMFAAAAAAAAAAVTDAEBAQEBAQEGSQICAAAAAwhJBgEBAQEBAQENUwUAAAAAAAAABSMkEgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARIMUwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUbCg4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQESWQAFAAAAAAAABQA5HQEBAQEBAQEGSQICAAAAAwgoFA4BAQEBAQETJQAFAAAAAAAABQVOEwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEaHzEFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBTkMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQETTgUAAAAAAAAABSBAGgEBAQEBARRICAICAAAAAwgESBQBAQEBAQEODSAFAAAAAAAAAAVgDAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDgwyIwUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUiLAwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMUwUAAAAAAAAABSYPDgEBAQEBAR44AgICAAAAAwMCRx4BAQEBAQEBDCMFAAAAAAAAAAULDwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMWVMABQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIBEBHQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEYIwUAAAAAAAAABUEPAQEBAQEBAQZWAgICAAAAAwQCRAYBAQEBAQEBDxEFAAAAAAAAAAU/IQ4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBEyE/BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIhcSAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR9OPwUAAAAAAAAABT0NAQEBAQEBAQY0AgICAAAAAwMHGQYBAQEBAQEBHT0FAAAAAAAAAAUAGxUBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDVMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBRsNAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ0xBQUAAAAAAAAFIDITAQEBAQEBDgYnAgICAAAAAwMDJy4aAQEBAQEBEjo/BQAAAAAAAAAFEAwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEfGiAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSMPDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR4RBQAAAAAAAAAFIxQOAQEBAQEBFDgEAgICAAAACAgIBykUAQEBAQEBAQoiBQAAAAAAAAAFIx4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEMIgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAURDwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDgwABQAAAAAAAAAFCw0BAQEBAQEBBlYCBwcHAAAAAgICAlYGAQEBAQEBAQ0QBQAAAAAAAAAFBRUSAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR05IAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU/LBIBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBE10FAAAAAAAAAAAFRh0BAQEBAQEBBk0HBAQEAAAAAgICAkkGAQEBAQEBAR8hBQAAAAAAAAAABRANAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARcjBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIh4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDUEFAAAAAAAAAAAAHgEBAQEBAQEBBgIIAwMDAAAAAgICAgIGGgEBAQEBAQEePwUAAAAAAAAABSIMDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBChAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABU4KAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQESOiMFAAAAAAAAAAUiDwEBAQEBAQEaLAIDAwMDAAAAAgICAgJXHwEBAQEBAQEeEQUAAAAAAAAABQAyDQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEaFCMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABTYYDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKPQUAAAAAAAAAAAUxDQEBAQEBAQEGRQIEAwMDAAAAAgICAgI0BgEBAQEBAQETOSAFAAAAAAAAAAURDwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEPCwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUQDwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEeQQUAAAAAAAAABT8hDgEBAQEBAQEULwMDAwMDAAAAAgICAgInXhQBAQEBAQEODSMFAAAAAAAAAAU/DA4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARM6AAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU/HRIBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR9MAAUAAAAAAAAABTYPAQEBAQEBARQpAgMDAwMDAAAAAgICAgICKRcBAQEBAQEBDBEFAAAAAAAAAAUFUQ0BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDg9PBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFUwwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQwRBQAAAAAAAAAABTcMAQEBAQEBAQZEAgQDAwMDAAAAAgICAgICSwYBAQEBAQEBH04FBQAAAAAAAAAFNgwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDzkFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBU4dAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBGiQ/BQAAAAAAAAAFIxQaAQEBAQEBAQYEBwMDAwMDAAAAAgICAgICCAYOAQEBAQEBDg9PBQAAAAAAAAAFAFsfAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEOISIFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABUEPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBCmAFAAAAAAAAAAAFCw0BAQEBAQEBFCoHAwQEBAQEAAAABAQEBAQECTAXAQEBAQEBAQ09BQUAAAAAAAAABUEPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEPEQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQA9DQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDCIFAAAAAAAAAAUAVRMBAQEBAQEBBjACAgICAgICAAAAAwMDAwMDB0MGAQEBAQEBARoYIAUAAAAAAAAABT8sEgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARoPIAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU/DAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEdOSAFAAAAAAAAAAVPDwEBAQEBAQEBHy8CAgICAgICAAAAAwMDAwMDCAdIFAEBAQEBAQEPEQUAAAAAAAAAAAURHgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQxTBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFEA8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEeQQUAAAAAAAAAAAVRCgEBAQEBAQEeRgICAgICAgICAAAAAwMDAwMDBAJSFwEBAQEBAQETVSYFAAAAAAAAAAAFGA4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHTkABQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFPywfAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARNOBQUAAAAAAAAABSMSEgEBAQEBAQEGRQICAgICAgICAAAAAwMDAwMDAwgZBgEBAQEBAQEBDAsFAAAAAAAAAAAFEQ8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHjYFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABREPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQw2BQAAAAAAAAAABREPAQEBAQEBARReBAICAgICAgICAAAAAwMDAwMDAwMJQhcBAQEBAQEBH0AFBQAAAAAAAAAFP04SAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKVAUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQBVEwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHSUgBQAAAAAAAAAFAA0aAQEBAQEBAQZWAgICAgICAgICAAAAAwMDAwMDAwQJTQYBAQEBAQEBAQ82BQAAAAAAAAAABVwPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARoTIwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU2Cg4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDyIFAAAAAAAAAAAFEQ8BAQEBAQEBAQYZAgICAgICAgICAAAAAwMDAwMDAwMDAgYSAQEBAQEBAR05AAUAAAAAAAAABQBZEwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAR4QBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFJQ0BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEVMgUFAAAAAAAAAAUgVR8BAQEBAQEBHjgCAgICAgICAgICAAAABAQEBAQEBAQEAkMGAQEBAQEBAQ4NIgUAAAAAAAAAAAU2DA4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHzo/BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIx8OAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARoVIgUAAAAAAAAAAAUQDAEBAQEBAQEBBhkCAgICAgICAgICAAAABwcHBwcHBwcHBy8BFAEBAQEBAQEdJQUAAAAAAAAAAAUFOhUBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDwsFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABTEMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQwbBQAAAAAAAAAABQVOEwEBAQEBAQEeOAcHCAgICAgICAgIAAAAAgICAgICAgICCAI8BgEBAQEBAQEBDVoFAAAAAAAAAAAFIxQOAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQETWwUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUhEgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBElkmBQAAAAAAAAAFBREMAQEBAQEBAQEGGQIHBAMDAwMDAwMDAAAAAgICAgICAgICCAgEBg4BAQEBAQEBEzogBQAAAAAAAAAABRAPAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4VNgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAULDwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBD1oFAAAAAAAAAAAFIAwSAQEBAQEBAR9OBwIHBAMDAwMDAwMDAAAAAgICAgICAgICCAQCVgYBAQEBAQEBAQ9TBQAAAAAAAAAABQUhHwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ81BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUAMg0BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEdTgUFAAAAAAAAAAUFEAwBAQEBAQEBAQZDAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAQIBBoUAQEBAQEBARJZPwUAAAAAAAAAAAUiDw4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBGgwmBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFNhgOAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARIkIwUAAAAAAAAAAAUiHRoBAQEBAQEBGjoCAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAQDAjAGAQEBAQEBAQEPUQUAAAAAAAAAAAUFNQwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDBAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBT0NAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ8RBQAAAAAAAAAAAAUyCgEBAQEBAQEBBkMCAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDCAQkFAEBAQEBAQEBHiMFAAAAAAAAAAAFAE4dAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEfWQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSMTEgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDFQFBQAAAAAAAAAABUEeAQEBAQEBAQEeRwgCAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDBAc8BgEBAQEBAQEBHTkgBQAAAAAAAAAABTYNDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEKCwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUQDAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEfJCMFAAAAAAAAAAAFP04TAQEBAQEBARoGGQICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAQDAwMCBhoBAQEBAQEBAQ9aBQAAAAAAAAAAAAVTDwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQo5AAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUgVR8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4MNgUAAAAAAAAAAAAFEAwBAQEBAQEBAR44AgICAgIHBAMDAwMDAwMDAAAACQkJCQkJCQkJCAgICAgCNAYBAQEBAQEBARofIwUAAAAAAAAAAAUgOQ0BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDg8jBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFQR4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ9TBQAAAAAAAAAAAAU2DQ4BAQEBAQEBGgEoAgkJCQkHCAgICAgICAgIAAAABAQEBAQEBAQEBwICAgICAjMXAQEBAQEBAQEMVAUFAAAAAAAAAAAFJgofAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDRsFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBSEfAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBD1UFBQAAAAAAAAAABQVVHQEBAQEBAQEBBkQCBAQEBAQIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAigGDgEBAQEBAQEBDxEFAAAAAAAAAAAABT8UHwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQESFCMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSMPDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEdTj8FAAAAAAAAAAAABTUMAQEBAQEBAQEUDgkIAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgJWBgEBAQEBAQEBHywjBQAAAAAAAAAABQVTHhoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEPUwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUbDAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARIPIwUAAAAAAAAAAAAFIg8OAQEBAQEBAQ4GQwIEAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgIILBQBAQEBAQEBAQ1OBQUAAAAAAAAAAAUFEQ0BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARNZAAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAU/DA4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDh5BBQAAAAAAAAAAAAUgIR8BAQEBAQEBAQZCAgMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICGQYBAQEBAQEBAQEPEAUFAAAAAAAAAAAFBRsMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDgw2BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFEQ8BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDTsFBQAAAAAAAAAABQA9DwEBAQEBAQEBFF4EBwQDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICAi0GAQEBAQEBAQEOGDYFAAAAAAAAAAAABQA5DAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHToFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFAFsfAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEPNQUFAAAAAAAAAAAABREMAQEBAQEBAQEBBksCBAMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICAgJIFAEBAQEBAQEBHx4/BQAAAAAAAAAAAAUAVQwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEOIT8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABUEMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQw9BQUAAAAAAAAAAAAFTw8OAQEBAQEBAQEGMwIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQkZBg4BAQEBAQEBAR1OIAUAAAAAAAAAAAAFPzIMAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEPOwUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQA5HQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEODxsFBQAAAAAAAAAAAAUmJBMBAQEBAQEBARReLwIIAwMDAwMDAwMIAgICAgICAgICAAAABAQEBAQEBAQEBwICAgICAgICCQQCVgYBAQEBAQEBAQEMVAAFAAAAAAAAAAAABSBVDAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARoKIAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUjHgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARoPPQUFAAAAAAAAAAAABSAlCgEBAQEBAQEBDgY0AgIIBAQEBAQEBAQIAgICAgICAgICAAAAAgICAgICAgICCAMDAwMDAwMDCAICB1AeAQEBAQEBAQEBDBAFBQAAAAAAAAAAAAUgMgwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQxTBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFXwwBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQwbBQUAAAAAAAAAAAAFBUYNAQEBAQEBAQEBBi0HAwMJAgICAgICAgIHAwMDAwMDAwMDAAAAAgICAgICAgICCAQDAwMDAwMDCAICAgQXGgEBAQEBAQEBDg9BBQUAAAAAAAAAAAAFP1UMAQEBAQEBAQEBAQEBAQEBAQEBAQEBCiUgBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIxQSAQEBAQEBAQEBAQEBAQEBAQEBAQEBDz0ABQAAAAAAAAAAAAUFER4BAQEBAQEBAQEeWQgIAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgI0BgEBAQEBAQEBAQEMQQUAAAAAAAAAAAAABQA5DAEBAQEBAQEBAQEBAQEBAQEBAQEODEEFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABVMPAQEBAQEBAQEBAQEBAQEBAQEBAQ4NNQUFAAAAAAAAAAAABQURDwEBAQEBAQEBARQPBAcEAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICPAYBAQEBAQEBAQESFSIFAAAAAAAAAAAAAAUAGwoOAQEBAQEBAQEBAQEBAQEBAQEVOQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSA6HwEBAQEBAQEBAQEBAQEBAQEBEh47BQUAAAAAAAAAAAAFBTYMEgEBAQEBAQEBGgZLBwQDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAisGAQEBAQEBAQEBEiEiBQAAAAAAAAAAAAAFBTsMEgEBAQEBAQEBAQEBAQEBAQ4PIwUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVBDQEBAQEBAQEBAQEBAQEBAQETD1MFBQAAAAAAAAAAAAAFNgwaAQEBAQEBAQEBBkQCAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgQ4HwEBAQEBAQEBARNOIgUAAAAAAAAAAAAABQVTFB8BAQEBAQEBAQEBAQEBAQ87BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFOQoBAQEBAQEBAQEBAQEBAR0sIgUFAAAAAAAAAAAAAAU2EhIBAQEBAQEBAQEGQgIDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgIWBhQBAQEBAQEBAQETISYFAAAAAAAAAAAAAAUFIgoNAQEBAQEBAQEBAQEBElkmBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIwwBAQEBAQEBAQEBAQEBCiE/BQUAAAAAAAAAAAAABSMPEgEBAQEBAQEBARcrBAgDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICKAYUAQEBAQEBAQEBEiEiBQAAAAAAAAAAAAAABSY5Dw4BAQEBAQEBAQEBHgsFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBTsKAQEBAQEBAQEBARIPMQAFAAAAAAAAAAAAAAAFNg8OAQEBAQEBAQEBHlknCQQDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAQEBAQEBAQECAICAgICAgICAjQGAQEBAQEBAQEBARNOIgUAAAAAAAAAAAAAAAUgEA0TAQEBAQEBAQEfHwUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSAeEgEBAQEBAQEBCh9aBQUAAAAAAAAAAAAABQU2EhIBAQEBAQEBAQEfBigCBAQEBAQEBAQJAgICAgICAgIHBAQEBAQEBAQEAAAABwcHBwcHBwcHBwcHBwcHBwcHBwcHBwcHBwcHBwI+BgEBAQEBAQEBAQETISIFBQAAAAAAAAAAAAAFBUEUDA4BAQEBAQENNgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAVTDAEBAQEBAQ4MVSMFBQAAAAAAAAAAAAAFBTYMEgEBAQEBAQEBAQEGKAIIBwcHBwcHBwcHBwcHBwcHBwcHBwcHBwcHBwcHAAAABAQEBAQEBAQEBwICAgICAgICCQQEBAQEBAQEBAICVh4BAQEBAQEBAQEBEhVBBQUAAAAAAAAAAAAAAAUAVB4SAQEBAQwbBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUmVRMBAQEBHQ81BQUAAAAAAAAAAAAAAAUFEQwaAQEBAQEBAQEBAQY0AgMCAgICAgICAgIIBAQEBAQEBAQIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAlIGAQEBAQEBAQEBARIMQQUFAAAAAAAAAAAAAAAFAE8hDw4BDh4/BQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFIh4OAQ4PTiIFBQAAAAAAAAAAAAAABQURDxIBAQEBAQEBAQEBBlYHBAMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgJSBg4BAQEBAQEBAQEBDxAABQAAAAAAAAAAAAAABQUgOxgTDxAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBVQMHRQRBQUFAAAAAAAAAAAAAAAFID0eAQEBAQEBAQEBARQGPgIDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgIJUgYBAQEBAQEBAQEBDgxUIAUAAAAAAAAAAAAAAAAFBSIsJD8FAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABSMPOiMFBQAAAAAAAAAAAAAAAAUmJQ0OAQEBAQEBAQEBAQZWBwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICCVIGAQEBAQEBAQEBAQEMTj8FBQAAAAAAAAAAAAAABQUAIgUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUiBQUFAAAAAAAAAAAAAAAFBU8kCgEBAQEBAQEBAQEOBj4HAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgJSHgEBAQEBAQEBAQEBHR42BQUAAAAAAAAAAAAAAAAFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBQAAAAAAAAAAAAAAAAUAER4TAQEBAQEBAQEBAQEGVgIDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICVgYBAQEBAQEBAQEBAR8YEAUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBSA9DA4BAQEBAQEBAQEBDgZWBwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAj4GFAEBAQEBAQEBAQEOD04jBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFIiEPAQEBAQEBAQEBAQEaBhkCBAMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwQIAgICAgICAgICAAAAAwMDAwMDAwMDBwICAgICAgICCQMDAwMDAwMDAwICAgICAgICAgI0BhQBAQEBAQEBAQEBARUsEQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBQU1FB8BAQEBAQEBAQEBARcuGQIDAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAgICAgICAgICCAQEBAQEBAQECAICAgICAgICAgMEBAQEBAQEBAcCKAYfAQEBAQEBAQEBAQEfD1QjBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFNlUPGgEBAQEBAQEBAQEBBisnAggCAgICAgICAgIEBAQEBAQEBAQJAgICAgICAgIHBAQEBAQEBAQEAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAhY4BgEBAQEBAQEBAQEBAQwfWiAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABT8QDQoBAQEBAQEBAQEBAQEGQgQJBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgIEKwYBAQEBAQEBAQEBAQESDDkjBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFNiUPDgEBAQEBAQEBAQEBDgYtAggEBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAjwGGgEBAQEBAQEBAQEBAR0eUT8FBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBSNRHhMBAQEBAQEBAQEBAQEUD0sCAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgI0Fx4BAQEBAQEBAQEBAQEBD1lTIAUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUgEBIKAQEBAQEBAQEBAQEBAQYULwIDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICBFAGDgEBAQEBAQEBAQEBARIPOloFBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFBVo6DxoBAQEBAQEBAQEBAQEaBi0HBwQDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgdWBhQBAQEBAQEBAQEBAQEBEw0lIgAFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBSALWQ0TAQEBAQEBAQEBAQEBARQuNAIIBAMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgICGUgGAQEBAQEBAQEBAQEBAQEdDTk2BQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUFUyUKHwEBAQEBAQEBAQEBAQEBBkgoAgMEAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAQEBAQEBAQECAICAgICAgICAgMEBAQEBAQEBAcCAgICAgICAgIECQItBh4BAQEBAQEBAQEBAQEBDh0PQAsmBQUFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABQUFPxENDxMBAQEBAQEBAQEBAQEBAR4GWAICAgMEBAQEBAQEBAcCAgICAgICAgIEBAQEBAQEBAQJAgICAgICAgIHBAQEBAQEBAQEAAAACAgICAgICAgIBwkJCQkJCQkJBwgICAgICAgICAkJCQkJCQkJCQcICAgICAgICAgJCQICKEYGDgEBAQEBAQEBAQEBAQEBHwxVESYFBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFACNRTgoaAQEBAQEBAQEBAQEBAQEUBl0vBwgICAkJCQkJCQkJCQcICAgICAgICAgJCQkJCQkJCQkICAgICAgICAgICQkJCQkJCQkJAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAghKBhcBAQEBAQEBAQEBAQEBAQETDyE9NgUFBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFBT9BPRgPEgEBAQEBAQEBAQEBAQEBDgYSQwgHBAMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIBAMDAwMDAwQIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICKDMGHwEBAQEBAQEBAQEBAQEBAQ4ND04RIyAFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFBT8jG1sMDQ4BAQEBAQEBAQEBAQEBAQEeBlIIAgMEAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgI0BgYUAQEBAQEBAQEBAQEBAQEBDh8PGDkRPwUFBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFBQUiEU4MDBIBAQEBAQEBAQEBAQEBAQEBHwYrGQcIBAMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgICAjwkBhQBAQEBAQEBAQEBAQEBAQEBDhMPHkAQIiMFBQUFBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFBQUFBSMiEAoeChIBAQEBAQEBAQEBAQEBAQEBARQGR00CAgQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwcEMBoGDgEBAQEBAQEBAQEBAQEBAQEBARINCjo9ESMgAAUFBQUAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAUFBQUAICMROSwMDRoBAQEBAQEBAQEBAQEBAQEBAQEaBl5FCAICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAgRWBgYUAQEBAQEBAQEBAQEBAQEBAQEBARIdDwwNJVNBIwUFBQUFBQUFBQUFBQAAAAAABQUFBQAAAAAFBQUFBQUFBQUFBQUjCxAlDQwPHRIBAQEBAQEBAQEBAQEBAQEBAQEBFAYGSgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwgCBDwBBhIBAQEBAQEBAQEBAQEBAQEBAQEBAQEOEw0YJE5OJTtPIz8/IAAABQUFBQUFBQUFBQUFBQUFAAAgPz8jTzslTk4kCg0TDgEBAQEBAQEBAQEBAQEBAQEBAQEBAR8GOEMIAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDAwMCAgICAgICAgIIAwMDAwMDAwMIAgICAgICAgICAAAAAwMDAwMDAwMEBwICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDBAMCAgICAgICAgIIBAMDBAQECAIvQwYGFwEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBEhIfHQwMDR1MPRsbMRAQEREREREREREREBAxGxsyTBUKDAwKExISAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEXBgY0CAICAgICAgICCQQDAwMDAwMDBAICAgICAgICAgcEAwMDAwMDBAMCAgICAgICAgIIBAMDAwMDAwQIAgICAgICAgICAAAACAgICAgICAgIBwcHBwcHBwcHBwgICAgICAgICAcHBwcHBwcHBwcICAgICAgICAgHBwcHBwcHBwcICAgICAgICAgHAgJNQgYGFAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4SHRUVFRgYDAwPDw8PDwwMGBgVFRUTEg4BAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQ4aBgZCGQICBwcHBwcHBwcHBwgICAgICAgICAcHBwcHBwcHBwcICAgICAgICAgHBwcHBwcHBwcICAgICAgICAgIBwcHBwcHBwcHAAAAAgICAgICAgICCAQEBAQEBAQECAICAgICAgICAgMEBAQEBAQEBAcCAgICAgICAgIEBAQEBAQEBAQJAgICAgICAgIHBAMJCS9ESAYXDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEaBgZMRS8JAgICCAQEBAQEBAQECAICAgICAgICAgMEBAQEBAQEBAcCAgICAgICAgIEBAQEBAQEBAQJAgICAgICAgIHBAQEBAQEBAQEAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMEAwgCB0MwBgYXFAEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEUBgYuMEkIAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMECAcJCEspXgYeGgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBHgYrLTQIAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMEAgICJzQqBgYGFBoBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBARIXBgYGMzQEAgcDAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMEAgICAgICAklWKS4GBh4UDgEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBFB4GBi4pSgQCAgkDBAMEAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMEAgICAgICAgICBycZREdIFAYGHhoOAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBDh8XBgYOSEdELycHCQgIBAQDAwMEAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMEAgICAgICAgICCAMHAgIEKElFMy4GBgYGBhcXFBQaEg4BAQEBAQEBAQEBAQ4SGhQUFwYGBgYGBkZCRTQvBAICAgIJBAMDAwMDAwMEAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMEAgICAgICAgICCAMDBAMICAgCBwcDBBlDREIpKissHAYGBgYGBgYGBgYGBgYcLCsqKS1EQxkECAgJAgICAgICAgIHBAMDAwMDAwMEAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMEAgICAgICAgICCAMDAwMDAwMECAICAgICAgICCAQnFigoGRkZGRkZGRkZKCgWJwQIAgkJBwcICAQJAgICAgICAgIHBAMDAwMDAwMEAgICAgICAgICCAMDAwMDAwMDCAICAgICAgICAgMDAwMDAwMDBAcCAgICAgICAgIEAwMDAwMDAwQJAgICAgICAgIHBAMDAwMDAwMDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"""

style = ttk.Style()
style.configure('W.TCombobox',font =('Courier New', '30'))

    
icondata= base64.b64decode(icon)
## The temp file is icon.ico
tempFile= "icon.ico"
iconfile= open(tempFile,"wb")
## Extract the icon
iconfile.write(icondata)
iconfile.close()
root.wm_iconbitmap(tempFile)
#top.wm_iconbitmap(tempFile)

## Delete the tempfile
os.remove(tempFile)

#top.title("CONFIGURATION")
root.title("RX2-REPORTER")
root.geometry('330x230+0+0')
root.call('wm', 'attributes', '.', '-topmost', True)
#root.resizable(False,False)
root.protocol('WM_DELETE_WINDOW', quitt)
root.mainloop()
